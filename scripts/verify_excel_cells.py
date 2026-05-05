"""
Verify all required cells in a generated feasibility Excel report.
Usage:
    python scripts/verify_excel_cells.py <path_to_xlsx>
    python scripts/verify_excel_cells.py <cloudinary_url>
"""

import os
import sys
import tempfile
import urllib.request

import openpyxl

# (cell, sheet, semantic_name, kind, required)
CELLS = [
    # Details sheet — black (service-sourced)
    ("A1", "Details", "report_header_title", "black", False),
    ("M1", "Details", "cts_fp_no_label", "black", False),
    ("M2", "Details", "village_name", "black", True),
    ("P4", "Details", "area_sqm", "black", True),
    ("P7", "Details", "setback_area_sqm", "black", False),
    ("N17", "Details", "reservation_area_sqm", "black", False),
    ("R17", "Details", "road_width_m", "black", True),
    ("R29", "Details", "noc_asi", "black", False),
    ("R30", "Details", "noc_mhcc", "black", False),
    ("R31", "Details", "noc_civil_aviation", "black", False),
    ("R32", "Details", "noc_crz", "black", False),
    ("O47", "Details", "existing_commercial_carpet_sqft", "black", False),
    ("Q47", "Details", "existing_residential_carpet_sqft", "black", False),
    ("J54", "Details", "rr_open_land_sqm", "black", True),
    # Details sheet — yellow (user/OCR inputs)
    ("N49", "Details", "num_commercial", "yellow", True),
    ("P49", "Details", "num_flats", "yellow", True),
    ("O51", "Details", "corpus_commercial", "yellow", False),
    ("Q51", "Details", "corpus_residential", "yellow", False),
    # P&L / financial
    ("D19", "Profit & Loss Statement", "sale_rate_commercial_gf", "financial", True),
    ("D28", "Profit & Loss Statement", "sale_rate_residential", "financial", True),
    ("D30", "Profit & Loss Statement", "parking_price_per_unit", "financial", True),
    # Construction Cost
    ("D8", "Construction Cost", "const_rate_commercial", "yellow", False),
    ("D12", "Construction Cost", "const_rate_residential", "yellow", False),
]

FALLBACK_SENTINELS = {
    "area_sqm": 0,
    "rr_open_land_sqm": 128870,
    "num_flats": 138,
    "num_commercial": 12,
}


def load_wb(path_or_url: str):
    if path_or_url.startswith("http"):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            urllib.request.urlretrieve(path_or_url, f.name)
            tmp = f.name
        wb = openpyxl.load_workbook(tmp, data_only=True)
        os.unlink(tmp)
    else:
        wb = openpyxl.load_workbook(path_or_url, data_only=True)
    return wb


def check(wb):
    results = []
    sheets = wb.sheetnames

    for cell_ref, sheet, name, kind, required in CELLS:
        if sheet not in sheets:
            results.append((cell_ref, sheet, name, kind, required, None, "SKIP", "sheet missing"))
            continue

        ws = wb[sheet]
        val = ws[cell_ref].value

        # Determine pass/fail
        is_empty = val is None or val in {"", 0}
        is_sentinel = val == FALLBACK_SENTINELS.get(name)

        if is_empty and required:
            status = "FAIL"
            note = "empty/zero — required"
        elif is_sentinel and required:
            status = "WARN"
            note = f"fallback sentinel {val!r}"
        elif is_empty:
            status = "WARN" if required else "OK"
            note = "empty/zero" if is_empty else ""
        else:
            status = "OK"
            note = ""

        results.append((cell_ref, sheet, name, kind, required, val, status, note))

    return results


def print_report(results):
    fail = [r for r in results if r[6] == "FAIL"]
    [r for r in results if r[6] == "WARN"]
    [r for r in results if r[6] == "OK"]
    [r for r in results if r[6] == "SKIP"]

    col = {"OK": "\033[92m", "WARN": "\033[93m", "FAIL": "\033[91m", "SKIP": "\033[90m"}

    for _cell_ref, _sheet, _name, _kind, _required, val, status, _note in results:
        str(val)[:30] if val is not None else "(none)"
        col.get(status, "")

    if fail:
        for _r in fail:
            pass
    else:
        pass

    return len(fail) == 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(1)

    target = sys.argv[1]
    wb = load_wb(target)
    results = check(wb)
    ok = print_report(results)
    sys.exit(0 if ok else 1)
