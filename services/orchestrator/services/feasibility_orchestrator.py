"""
Feasibility Analysis Orchestrator Service
Orchestrates all microservices for full feasibility analysis.
"""

import asyncio
import contextlib
import logging
import os
from collections.abc import Coroutine
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from dhara_shared.services.cache import redis_cache
from dhara_shared.services.http import AsyncHTTPClient
from fastapi import BackgroundTasks
from sqlalchemy.orm.attributes import flag_modified

from ..core.circuit_breaker import CircuitOpenError, call_with_circuit_breaker
from ..core.config import settings
from ..db import async_session_factory
from ..models import FeasibilityReport, ReportStatus
from .cloudinary import upload_content
from .dossier_service import dossier_service

logger = logging.getLogger(__name__)

# Where to write generated reports (inside the shared volume)
_REPORTS_DIR = Path(os.getenv("REPORTS_DIR", Path(__file__).parent.parent / "generated_reports"))
_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Redis instance for report path tracking
from dhara_shared.services.cache import RedisCache

_cache = RedisCache()
REPORT_PATH_PREFIX = "report_path"
REPORT_PATH_TTL = 3600  # 1 hour


def get_report_path(job_id: str) -> str | None:
    """Get report path from Redis."""
    return _cache.get(f"{REPORT_PATH_PREFIX}:{job_id}")


def set_report_path(job_id: str, path: str):
    """Set report path in Redis."""
    _cache.set(f"{REPORT_PATH_PREFIX}:{job_id}", path, ttl=REPORT_PATH_TTL)


MCGM_URL = settings.MCGM_PROPERTY_URL
SITE_ANALYSIS_URL = settings.SITE_ANALYSIS_URL
DP_REMARKS_URL = settings.DP_REPORT_URL
PR_CARD_URL = settings.PR_CARD_URL
AVIATION_HEIGHT_URL = settings.HEIGHT_URL
READY_RECKONER_URL = settings.READY_RECKONER_URL
REPORT_GENERATOR_URL = settings.REPORT_URL
OCR_URL = settings.OCR_URL


def _rr_locality_from_req(req: dict) -> tuple[str, str]:
    """Derive (rr_locality, taluka) from req village + ward for Ready Reckoner lookup."""
    village = (req.get("village") or "").strip()
    ward = (req.get("ward") or "").upper()

    # Normalize: lowercase, remove spaces/hyphens/dots
    key = village.lower().replace(" ", "").replace("-", "").replace(".", "")

    # Disambiguate Vile Parle by ward: K/W = West, K/E = East
    if key == "vileparle":
        if "W" in ward:
            return "vile-parle-west", "andheri"
        elif "E" in ward:
            return "vile-parle-east", "andheri"

    # Common village -> RR locality mappings (high-frequency Mumbai areas)
    # RR service has full data, this is just optimization to avoid API calls
    village_to_rr = {
        "andheri": ("andheri", "andheri"),
        "bandra": ("bandra-a", "andheri"),
        "kurla": ("kurla", "kurla"),
        "dharavi": ("dharavi", "kurla"),
        "borivali": ("borivali", "borivali"),
        "malad": ("malad", "borivali"),
        "kandivali": ("kandivali", "borivali"),
        "goregaon": ("goregaon", "borivali"),
        "santacruz": ("santacruz", "andheri"),
        "khar": ("khar", "andheri"),
        "juhu": ("juhu", "andheri"),
    }

    if key in village_to_rr:
        return village_to_rr[key]

    # Generic slug fallback: derive from address
    slug = village.lower().replace(" ", "-").replace("/", "-")
    taluka = "andheri" if ward and ward[0] in ("H", "K", "P", "R", "S") else "mumbai-city"
    return slug or "bhuleshwar", taluka


class FeasibilityOrchestrator:
    """Orchestrates all microservices for feasibility analysis."""

    async def _update_buffer(self, job_id: str, key: str, data: dict):
        """Update the data_buffer in the DB incrementally."""
        try:
            async with async_session_factory() as db:
                report = await db.get(FeasibilityReport, job_id)
                if report:
                    buffer = report.data_buffer or {}
                    buffer[key] = data
                    report.data_buffer = buffer
                    # Also update output_data for backward compatibility during processing
                    report.output_data = {**(report.output_data or {}), **buffer}
                    flag_modified(report, "data_buffer")
                    flag_modified(report, "output_data")
                    await db.commit()
                    logger.debug(f"[{job_id}] Updated buffer key: {key}")
        except Exception as e:
            logger.warning(f"[{job_id}] Failed to update buffer: {e}")

    async def _refresh_report(self, job_id: str, req: dict):
        """Re-generate the report based on current data_buffer and update DB."""
        try:
            async with async_session_factory() as db:
                report = await db.get(FeasibilityReport, job_id)
                if not report:
                    return

                buffer = report.data_buffer or {}

                # Map buffer back to round1/round2 structure for call_report_generator
                round1 = {
                    "pr_card": buffer.get("pr_card", {}),
                    "mcgm": buffer.get("mcgm", {}),
                    "site_analysis": buffer.get("site_analysis", {}),
                    "dp_remarks": buffer.get("dp_remarks", {}),
                    "ocr": buffer.get("ocr", {}),
                    "tenements_ocr": buffer.get("tenements_ocr", {}),
                }
                round2 = {
                    "aviation_height": buffer.get("aviation_height", {}),
                    "ready_reckoner": buffer.get("ready_reckoner", {}),
                }

                async with AsyncHTTPClient(timeout=300.0, request_id=job_id) as client:
                    report_bytes = await self.call_report_generator(
                        client, round1, round2, req, job_id=job_id
                    )

                report_filename = f"feasibility_{job_id}.xlsx"
                report_path = str(_REPORTS_DIR / report_filename)
                Path(report_path).write_bytes(report_bytes)

                report.report_path = report_path
                set_report_path(job_id, report_path)
                await db.commit()
                logger.info(f"[{job_id}] Report refreshed at {report_path}")
        except Exception as e:
            logger.warning(f"[{job_id}] Failed to refresh report: {e}")

    async def run_round1(self, req: dict, job_id: str | None = None) -> dict:
        """Call Round 1 services and update buffer incrementally."""
        async with AsyncHTTPClient(timeout=300.0, request_id=job_id) as client:
            tasks = [
                ("pr_card", self.call_pr_card(client, req)),
                ("mcgm", self.call_mcgm(client, req)),
                ("site_analysis", self.call_site_analysis(client, req)),
                ("dp_remarks", self.call_dp_remarks(client, req)),
            ]

            import base64

            ocr_pdf_bytes = req.get("ocr_pdf_bytes")
            if not ocr_pdf_bytes and req.get("ocr_pdf_b64"):
                ocr_pdf_bytes = base64.b64decode(req["ocr_pdf_b64"])
            if ocr_pdf_bytes:
                tasks.append(
                    (
                        "ocr",
                        self.call_ocr_service(
                            client,
                            ocr_pdf_bytes,
                            doc_type="old_plan",
                            filename="old_plan.pdf",
                            content_type="application/pdf",
                        ),
                    )
                )

            tenements_pdf_bytes = None
            if req.get("tenements_sheet_b64"):
                tenements_pdf_bytes = base64.b64decode(req["tenements_sheet_b64"])
                tasks.append(
                    (
                        "tenements_ocr",
                        self.call_ocr_service(
                            client,
                            tenements_pdf_bytes,
                            doc_type="tenements_sheet",
                            filename=req.get("tenements_sheet_filename", "tenements.pdf"),
                            content_type=req.get("tenements_sheet_content_type", "application/pdf"),
                        ),
                    )
                )

            return await self._run_parallel(tasks, req, job_id=job_id)

    async def run_round2(
        self,
        lat: float | None,
        lng: float | None,
        zone: str | None,
        req: dict | None = None,
        job_id: str | None = None,
    ) -> dict:
        """Call Round 2 dependent services and update buffer incrementally."""
        req = req or {}
        async with AsyncHTTPClient(timeout=300.0, request_id=job_id) as client:
            tasks = []

            # Aviation Height needs lat/lng
            if lat and lng:
                tasks.append(("aviation_height", self.call_aviation_height(client, lat, lng)))

            # Ready Reckoner needs zone + req context (Fallback to 'Residential' if missing)
            effective_zone = zone or "Residential"
            tasks.append(("ready_reckoner", self.call_ready_reckoner(client, effective_zone, req)))

            return await self._run_parallel(tasks, req, job_id=job_id)

    async def _run_parallel(
        self,
        tasks: list[tuple[str, asyncio.Task | Coroutine]],
        req: dict,
        job_id: str | None = None,
    ) -> dict:
        """Run tasks in parallel, updating buffer and refreshing report after each."""
        results_out = {}
        pending = {asyncio.create_task(t[1]): t[0] for t in tasks}

        while pending:
            done, _ = await asyncio.wait(pending.keys(), return_when=asyncio.FIRST_COMPLETED)
            for task in done:
                name = pending.pop(task)
                try:
                    res = await task
                    results_out[name] = res
                    if job_id:
                        await self._update_buffer(job_id, name, res)
                        from .dossier_service import dossier_service

                        # Preserve progress tracking by using round1_ prefix for round 1 tasks
                        # Round 2 tasks go into 'round2' eventually, but here we can just update
                        stage_name = (
                            f"round1_{name}"
                            if name
                            in (
                                "mcgm",
                                "dp_remarks",
                                "pr_card",
                                "ocr",
                                "tenements_ocr",
                                "site_analysis",
                            )
                            else name
                        )
                        await dossier_service.update_dossier(job_id, stage_name, res)
                        await self._refresh_report(job_id, req)
                except Exception as e:
                    logger.warning(f"[{job_id}] Task {name} failed: {e}")
                    results_out[name] = {"error": str(e)}
        return results_out

    async def analyze(
        self,
        req: dict,
        background_tasks: BackgroundTasks | None = None,
        user_id: str | None = None,
        report_id: str | None = None,
    ) -> dict:
        """
        Main entry point for feasibility analysis.
        """
        # Step 0: Normalize request fields to prevent 'None' strings or type errors
        for k in ["ward", "village", "cts_no", "fp_no", "tps_name"]:
            val = req.get(k)
            if val is None:
                req[k] = ""
            else:
                req[k] = str(val).strip()

        job_id = report_id or str(uuid4())
        logger.info(
            f"[{job_id}] Starting feasibility analysis for society_id={req.get('society_id')}"
        )

        # Step 0: Ensure a DB record exists if we have a user_id
        if user_id:
            try:
                async with async_session_factory() as db:
                    report = await db.get(FeasibilityReport, job_id)
                    if not report:
                        new_report = FeasibilityReport(
                            id=job_id,
                            user_id=user_id,
                            society_id=req.get("society_id"),
                            title=req.get(
                                "title", f"Analysis: {req.get('society_name', 'Unnamed')}"
                            ),
                            status=ReportStatus.PROCESSING,
                            input_data=req,
                            data_buffer=req.get("manual_inputs", {}),  # Start with manual inputs
                        )
                        db.add(new_report)

                        # NEW: Also create a stub in society_reports so FE history/counts work
                        from ..models.report import SocietyReport

                        soc_report = SocietyReport(
                            id=job_id,  # Use same ID for perfect linking
                            society_id=req.get("society_id"),
                            title=new_report.title,
                            report_type="feasibility",
                            status=ReportStatus.PROCESSING,
                        )
                        db.add(soc_report)

                        await db.commit()

                # Persist manual_inputs to DB buffer so they are available for all refreshes
                await self._update_buffer(job_id, "manual_inputs", req)

                # Initial refresh with just manual inputs
                await self._refresh_report(job_id, req)
            except Exception as e:
                logger.warning(f"[{job_id}] Could not create initial report record: {e}")

        # Step 0a: If ward/village empty, resolve from address FIRST to provide context for CTS/FP
        if not req.get("ward") or not req.get("village"):
            address = req.get("address")

            # Fetch address from society if not provided in req
            if not address and req.get("society_id"):
                try:
                    from ..models.society import Society

                    async with async_session_factory() as db:
                        soc = await db.get(Society, req.get("society_id"))
                        if soc and soc.address:
                            address = soc.address
                except Exception as e:
                    logger.warning(f"[{job_id}] Failed to fetch society address: {e}")

            if not address:
                address = req.get("society_name", "")
            if address:
                try:
                    from .address_resolver import resolve_address_from_input

                    resolved = await resolve_address_from_input(address)
                    if resolved:
                        if resolved.get("ward") and not req.get("ward"):
                            req["ward"] = resolved["ward"]
                        if resolved.get("village") and not req.get("village"):
                            req["village"] = resolved["village"]
                        if resolved.get("taluka") and not req.get("taluka"):
                            req["taluka"] = resolved["taluka"]
                        if resolved.get("district") and not req.get("district"):
                            req["district"] = resolved["district"]
                        logger.info(
                            f"[{job_id}] Address resolved: ward={req.get('ward')} village={req.get('village')}"
                        )
                except Exception as addr_err:
                    logger.warning(f"[{job_id}] Address resolution failed: {addr_err}")

        # Step 0b: Resolve CTS/FP using ward/village context if available
        cts_no = req.get("cts_no")
        fp_no = req.get("fp_no")
        if cts_no or fp_no:
            from .cts_fp_resolver import get_resolver

            resolver = get_resolver()
            res = await resolver.resolve(
                cts_no=cts_no,
                fp_no=fp_no,
                village=req.get("village"),
                ward=req.get("ward"),
                address=req.get("address"),
            )
            if res:
                if res.cts_no:
                    req["cts_no"] = res.cts_no
                if res.fp_no:
                    req["fp_no"] = res.fp_no
                if res.tps_name:
                    req["tps_name"] = res.tps_name
                # ArcGIS location data is authoritative
                if res.extra:
                    ext = res.extra
                    if not req.get("ward"):
                        req["ward"] = ext.get("ward")
                    if not req.get("village"):
                        req["village"] = ext.get("village")
                    if not req.get("taluka"):
                        req["taluka"] = ext.get("taluka")
                    if not req.get("district"):
                        req["district"] = ext.get("district")
                    if not req.get("plot_area_sqm") and ext.get("area_sqm"):
                        req["plot_area_sqm"] = float(ext["area_sqm"])
                logger.info(f"[{job_id}] CTS/FP resolved: {req.get('cts_no')} / {req.get('fp_no')}")

                # Update buffer with resolved land details immediately
                await self._update_buffer(
                    job_id,
                    "resolved_identifiers",
                    {
                        "cts_no": req.get("cts_no"),
                        "fp_no": req.get("fp_no"),
                        "tps_name": req.get("tps_name"),
                        "ward": req.get("ward"),
                        "village": req.get("village"),
                    },
                )
                await self._refresh_report(job_id, req)

                # All gathered land identity data (CTS, Plot Area, Ward) is now stored
                # exclusively in the FeasibilityReport record and the data_buffer.
                # We NO LONGER pollute the base Society Master Record.

        # Round 1: parallel calls with incremental buffer updates
        round1_results = await self.run_round1(req, job_id=job_id)
        await dossier_service.update_dossier(job_id, "round1", round1_results)

        # Extract dependencies for Round 2
        mc_result = round1_results.get("mcgm", {})
        sa_result = round1_results.get("site_analysis", {})

        lat = (
            req.get("lat")
            or mc_result.get("centroid_lat")
            or mc_result.get("lat")
            or sa_result.get("lat")
        )
        lng = (
            req.get("lng")
            or mc_result.get("centroid_lng")
            or mc_result.get("lng")
            or sa_result.get("lng")
        )

        zone = round1_results.get("dp_remarks", {}).get("zone_code") or round1_results.get(
            "dp_remarks", {}
        ).get("zone")

        logger.info(f"[{job_id}] Round 1 extracted: lat={lat}, lng={lng}, zone={zone}")

        # Enrich dp_remarks with GIS metrics already computed by mcgm service
        dp_result = round1_results.get("dp_remarks", {})
        if not dp_result.get("setback_area_sqm") and mc_result.get("setback_area_m2"):
            dp_result["setback_area_sqm"] = mc_result["setback_area_m2"]
            logger.info(f"[{job_id}] Setback from mcgm GIS: {dp_result['setback_area_sqm']} m²")
        if not dp_result.get("reservation_area_sqm") and mc_result.get("reservation_area_m2"):
            dp_result["reservation_area_sqm"] = mc_result["reservation_area_m2"]
            logger.info(
                f"[{job_id}] Reservation area from mcgm GIS: {dp_result['reservation_area_sqm']} m²"
            )
        round1_results["dp_remarks"] = dp_result

        # Round 2: dependent services with incremental updates
        round2_results = await self.run_round2(lat, lng, zone, req, job_id=job_id)
        await dossier_service.update_dossier(job_id, "round2", round2_results)

        # Round 3: Final upload to Cloudinary if needed
        # Get report path from DB (persisted via _refresh_report)
        report_path = None
        report_url = None
        report_error = None

        try:
            async with async_session_factory() as db:
                report = await db.get(FeasibilityReport, job_id)
                if report:
                    report_path = report.report_path
                    report_url = report.file_url
        except Exception as e:
            logger.warning(f"[{job_id}] Failed to fetch report path from DB: {e}")

        # Pro-level production filename: Society_Name_Scheme_Date_JobId.xlsx
        society_name = req.get("society_name", "Report")
        clean_name = "".join(c if c.isalnum() else "_" for c in society_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        scheme_label = (
            req.get("scheme", "33_7_B").replace("(", "").replace(")", "").replace(".", "_")
        )
        pro_filename = f"{clean_name}_{scheme_label}_{timestamp}_{job_id[:8]}.xlsx"

        # Fallback: report not generated via incremental refreshes (e.g. missing DB record)
        if not report_path or not os.path.exists(report_path):
            try:
                async with AsyncHTTPClient(timeout=300.0, request_id=job_id) as client:
                    report_bytes = await self.call_report_generator(
                        client, round1_results, round2_results, req, job_id=job_id
                    )
                if report_bytes:
                    report_path = str(_REPORTS_DIR / f"feasibility_{job_id}.xlsx")
                    # Use a safer write with potential chmod
                    try:
                        p = Path(report_path)
                        p.write_bytes(report_bytes)
                        with contextlib.suppress(Exception):
                            p.chmod(0o777)
                        logger.info(f"[{job_id}] Report generated via fallback path: {report_path}")

                        # Update DB with report path
                        try:
                            async with async_session_factory() as db:
                                report = await db.get(FeasibilityReport, job_id)
                                if report:
                                    report.report_path = report_path
                                    await db.commit()
                        except Exception as db_err:
                            logger.warning(f"[{job_id}] Failed to update report path in DB: {db_err}")
                    except Exception as we:
                        logger.exception(f"[{job_id}] Failed to write report file: {we}")
            except Exception as e:
                logger.warning(f"[{job_id}] Fallback report generation failed: {e}")

        if report_path and os.path.exists(report_path):
            # Check if we should upload to cloud
            if not getattr(settings, "SAVE_REPORTS_LOCALLY", True):
                try:
                    report_bytes = Path(report_path).read_bytes()
                    upload_res = await upload_content(
                        content=report_bytes,
                        filename=pro_filename,  # Use professional filename for Cloudinary
                        folder="dhara/reports",
                        resource_type="raw",
                    )
                    report_url = upload_res.get("secure_url")
                    logger.info(
                        f"[{job_id}] Production report uploaded to Cloudinary: {report_url}"
                    )
                except Exception as ue:
                    logger.warning(f"[{job_id}] Final Cloudinary upload failed: {ue}")
        else:
            report_error = "Report file missing after rounds"

        final_result = {
            "job_id": job_id,
            "status": "completed" if not report_error else "failed",
            "round1_results": round1_results,
            "round2_results": round2_results,
            "report_generated": report_path is not None,
            "report_path": report_path,
            "report_url": report_url,
            "report_error": report_error,
        }

        await dossier_service.update_dossier(job_id, "final_result", final_result)

        # Update DB with final results and summary metrics
        try:
            async with async_session_factory() as db:
                report = await db.get(FeasibilityReport, job_id)
                if report:
                    report.status = (
                        ReportStatus.COMPLETED if not report_error else ReportStatus.FAILED
                    )
                    report.report_path = report_path
                    report.file_url = report_url
                    report.output_data = final_result

                    # ── Populate summary columns for dashboard visibility ─────
                    # Priority: user input > resolved land identity > mcgm result
                    report.ward = req.get("ward") or round1_results.get("mcgm", {}).get("ward")
                    report.village = req.get("village") or round1_results.get("mcgm", {}).get(
                        "village"
                    )
                    report.cts_no = req.get("cts_no") or round1_results.get("mcgm", {}).get(
                        "cts_no"
                    )
                    report.fp_no = req.get("fp_no") or round1_results.get("mcgm", {}).get("fp_no")

                    report.plot_area = req.get("plot_area_sqm") or round1_results.get(
                        "mcgm", {}
                    ).get("area_sqm")
                    report.fsi = req.get("fsi") or 2.7

                    grand_total = (
                        round2_results.get("ready_reckoner", {})
                        .get("rr_data", {})
                        .get("grand_total")
                    )
                    if grand_total:
                        report.estimated_value = str(grand_total)

                    if report_error:
                        report.error_message = report_error

                    # ── Sync to SocietyReport history stub ────────────────────
                    from ..models.report import SocietyReport

                    # Try to find by ID (if we started with same job_id) or title fallback
                    soc_report = await db.get(SocietyReport, job_id)
                    if not soc_report:
                        from sqlalchemy import select

                        stmt = (
                            select(SocietyReport)
                            .where(SocietyReport.title == report.title)
                            .order_by(SocietyReport.created_at.desc())
                            .limit(1)
                        )
                        res = await db.execute(stmt)
                        soc_report = res.scalar()

                    if soc_report:
                        soc_report.status = report.status
                        soc_report.file_url = report.file_url

                    await db.commit()
                    logger.info(f"[{job_id}] Finalized report DB records (status={report.status})")

        except Exception as e:
            logger.warning(f"[{job_id}] Could not finalize report record: {e}")

        if background_tasks:
            background_tasks.add_task(self._check_expiries, job_id, req)

        return final_result

    async def _call_async_service_poll(
        self,
        client: AsyncHTTPClient,
        service_url: str,
        endpoint: str,
        payload: dict | None = None,
        files: dict | None = None,
        timeout: float = 300.0,
        unwrap_result: bool = False,
    ) -> dict:
        """
        Unified helper for async services following Submit -> Poll pattern.
        """
        import time

        import httpx

        start = time.perf_counter()
        service_name = service_url.split("/")[-1] or "service"

        try:
            # 1. Submit Job
            if files:
                # Use raw httpx for multipart uploads
                async with httpx.AsyncClient(timeout=httpx.Timeout(timeout)) as raw_client:
                    resp = await raw_client.post(
                        f"{service_url}{endpoint}",
                        files=files,
                        data=payload,
                    )
            else:
                resp = await client.post(f"{service_url}{endpoint}", json=payload)

            resp.raise_for_status()
            job_data = resp.json()

            # Immediate completion (cache hit or fast path)
            if job_data.get("status") == "completed":
                data = job_data.get("result") if unwrap_result else job_data
                return data or {}

            # 2. Poll Status
            job_id = job_data.get("id")
            if not job_id:
                return {"error": f"No job ID returned from {service_name}"}

            data = await self._poll_service_status(client, service_url, job_id, timeout=timeout)

            # 3. Handle Result
            if unwrap_result:
                final_result = data.get("result", {}) or {}
                if data.get("error"):
                    final_result["error"] = data["error"]
                return final_result

            return data
        except Exception as e:
            duration = time.perf_counter() - start
            logger.warning(f"[PERF] {service_name} {endpoint} FAILED after {duration:.2f}s | Error: {e}")
            return {"error": str(e)}

    async def _poll_service_status(
        self,
        client: AsyncHTTPClient,
        service_url: str,
        job_id: str,
        timeout: float = 300.0,
    ) -> dict:
        """Poll a service's /status/{job_id} endpoint until completion."""
        import time

        start = time.perf_counter()
        poll_interval = 2.0  # seconds between polls
        while time.perf_counter() - start < timeout:
            try:
                resp = await client.get(f"{service_url}/status/{job_id}")
                resp.raise_for_status()
                data = resp.json()
                status = data.get("status", "").lower()
                if status in ("completed", "failed"):
                    return data
                await asyncio.sleep(poll_interval)
            except Exception as e:
                logger.warning(f"Polling failed for {service_url}/status/{job_id}: {e}")
                await asyncio.sleep(poll_interval)

        error_msg = f"Timeout waiting for {service_url}/status/{job_id} after {timeout}s"
        logger.warning(error_msg)
        return {"id": job_id, "status": "failed", "error": error_msg}

    async def call_pr_card(self, client: AsyncHTTPClient, req: dict) -> dict:
        """Call PR Card Scraper service."""
        payload = {
            "district": (req.get("district") or "Mumbai").strip(),
            "taluka": (req.get("taluka") or "Mumbai").strip(),
            "village": (req.get("village") or "").replace("-WEST", "").replace("-EAST", "").strip(),
            "survey_no": (req.get("cts_no") or req.get("survey_no") or req.get("fp_no") or "").strip(),
        }
        return await self._call_async_service_poll(client, PR_CARD_URL, "/scrape", payload)

    async def call_mcgm(self, client: AsyncHTTPClient, req: dict) -> dict:
        """Call MCGM Property Lookup service."""
        cts_no = req.get("cts_no") or req.get("fp_no", "")
        payload = {
            "ward": str(req.get("ward") or "M/E"),
            "village": str(req.get("village") or "").replace("-WEST", "").replace("-EAST", "").strip(),
            "cts_no": str(cts_no) if cts_no else "",
            "use_fp": bool(req.get("use_fp_scheme", False)),
            "tps_name": str(req.get("tps_name")) if req.get("tps_name") else None,
        }
        return await self._call_async_service_poll(client, MCGM_URL, "/lookup", payload)

    async def call_site_analysis(self, client: AsyncHTTPClient, req: dict) -> dict:
        """Call Site Analysis service."""
        import time

        start = time.perf_counter()
        try:
            resp = await client.post(
                f"{SITE_ANALYSIS_URL}/analyse",
                json={
                    "address": req.get("address", ""),
                    "ward": req.get("ward"),
                },
            )
            resp.raise_for_status()
            data = resp.json()
            duration = time.perf_counter() - start
            logger.info(f"[PERF] Site Analysis took {duration:.2f}s | Result: {str(data)[:200]}...")
            return data
        except Exception as e:
            duration = time.perf_counter() - start
            logger.warning(f"[PERF] Site Analysis FAILED after {duration:.2f}s | Error: {e}")
            return {"error": str(e)}

    async def call_dp_remarks(self, client: AsyncHTTPClient, req: dict) -> dict:
        """Call DP Remarks service."""
        import base64
        import time

        import httpx

        start = time.perf_counter()

        # ── Strategy 1: PDF upload (preferred for testing) ─────────────────
        dp_pdf_b64 = req.get("dp_remark_pdf_b64")
        if dp_pdf_b64:
            try:
                pdf_bytes = base64.b64decode(dp_pdf_b64)
                async with httpx.AsyncClient(timeout=120.0) as raw_client:
                    resp = await raw_client.post(
                        f"{DP_REMARKS_URL}/fetch/from-pdf",
                        files={"file": ("dp_remark.pdf", pdf_bytes, "application/pdf")},
                        data={
                            "ward": req.get("ward", ""),
                            "village": req.get("village", ""),
                            "cts_no": req.get("cts_no", ""),
                        },
                    )
                resp.raise_for_status()
                result = resp.json()
                duration = time.perf_counter() - start
                logger.info(
                    f"[PERF] DP Remarks (PDF) took {duration:.2f}s | Result: {str(result)[:200]}..."
                )
                return result
            except Exception as e_pdf:
                logger.warning(f"DP Remarks PDF parse failed ({e_pdf}); trying web fallback")

        # ── Strategy 2 & 3: Web scraper (Async + Polling) ──────────────────
        cts_no = req.get("cts_no")
        fp_no = req.get("fp_no", "")
        base_payload = {
            "ward": str(req.get("ward") or "M/E"),
            "village": str(req.get("village") or ""),
            "lat": float(req.get("lat")) if req.get("lat") is not None else None,
            "lng": float(req.get("lng")) if req.get("lng") is not None else None,
        }

        try:
            result = await self._call_async_service_poll(
                client,
                DP_REMARKS_URL,
                "/fetch",
                payload={
                    **base_payload,
                    "cts_no": str(cts_no or fp_no or ""),
                    "use_fp_scheme": bool(req.get("use_fp_scheme", False)),
                },
            )
            if result.get("cts_not_found") and fp_no:
                raise ValueError("CTS not in dropdown — retrying with FP")

            duration = time.perf_counter() - start
            logger.info(f"[PERF] DP Remarks (Web Async) took {duration:.2f}s")
            return result
        except Exception as e_cts:
            logger.warning(f"DP Remarks CTS attempt failed ({e_cts}); trying FP fallback")
            try:
                result = await self._call_async_service_poll(
                    client,
                    DP_REMARKS_URL,
                    "/fetch",
                    payload={
                        **base_payload,
                        "cts_no": fp_no,
                        "use_fp_scheme": True,
                        "tps_name": req.get("tps_name"),
                    },
                )
                duration = time.perf_counter() - start
                logger.info(f"[PERF] DP Remarks (Web FP Async) took {duration:.2f}s")
                return result
            except Exception as e_fp:
                duration = time.perf_counter() - start
                logger.warning(f"[PERF] DP Remarks FAILED after {duration:.2f}s | Error: {e_fp}")
                return {"error": str(e_fp)}

    async def call_aviation_height(self, client: AsyncHTTPClient, lat: float, lng: float) -> dict:
        """Call Aviation Height service."""
        import time

        start = time.perf_counter()
        try:
            resp = await client.post(
                f"{AVIATION_HEIGHT_URL}/check-height", json={"lat": lat, "lng": lng}
            )
            resp.raise_for_status()
            data = resp.json()
            duration = time.perf_counter() - start
            logger.info(
                f"[PERF] Aviation Height took {duration:.2f}s | Result: {str(data)[:200]}..."
            )
            return data
        except Exception as e:
            duration = time.perf_counter() - start
            logger.warning(f"[PERF] Aviation Height FAILED after {duration:.2f}s | Error: {e}")
            return {"error": str(e)}

    async def call_ready_reckoner(
        self, client: AsyncHTTPClient, zone: str, req: dict | None = None
    ) -> dict:
        """Call Ready Reckoner service."""
        import time

        start = time.perf_counter()
        req = req or {}
        try:
            locality, taluka = _rr_locality_from_req(req)
            payload = {
                "zone": str(zone),
                "locality": locality,
                "district": "mumbai",
                "taluka": taluka,
                "plot_area_sqm": req.get("plot_area_sqm") or 0.0,
                "property_type": "residential",
                "property_area_sqm": req.get("plot_area_sqm") or 0.0,
            }
            resp = await client.post(
                f"{READY_RECKONER_URL}/calculate",
                json=payload,
            )
            resp.raise_for_status()
            raw = resp.json()
            data = raw.get("data") or {}
            rr_rates = data.get("rr_rates", [])
            rates = {r["category"].lower(): r["value"] for r in rr_rates}
            rr_open = rates.get("land") or rates.get("open land") or 0
            rr_res = rates.get("residential") or rr_open
            rr_shop = rates.get("shop") or rates.get("office") or rr_open * 1.5

            result = {
                "rr_data": data,
                "rr_open_land_sqm": rr_open,
                "rr_residential_sqm": rr_res,
                "rr_shop_sqm": rr_shop,
                "sale_rate_residential": round(rr_res / 10.764 * 1.4) if rr_res else 0,
                "sale_rate_commercial_gf": round(rr_shop / 10.764 * 1.4) if rr_shop else 0,
                "parking_price_per_unit": 1_200_000,
            }
            duration = time.perf_counter() - start
            logger.info(
                f"[PERF] Ready Reckoner took {duration:.2f}s | Result: {str(result)[:200]}..."
            )
            return result
        except Exception as e:
            duration = time.perf_counter() - start
            logger.warning(f"[PERF] Ready Reckoner FAILED after {duration:.2f}s | Error: {e}")
            return {"error": str(e)}

    async def call_ocr_service(
        self,
        client: AsyncHTTPClient,
        file_bytes: bytes,
        doc_type: str = "old_plan",
        filename: str = "document.pdf",
        content_type: str = "application/pdf",
    ) -> dict:
        """Call OCR service."""
        return await self._call_async_service_poll(
            client,
            OCR_URL,
            "/extract",
            payload={"doc_type": doc_type},
            files={"file": (filename, file_bytes, content_type)},
            unwrap_result=True,
        )

    async def call_report_generator(
        self,
        client: AsyncHTTPClient,
        round1: dict,
        round2: dict,
        req: dict,
        job_id: str | None = None,
    ) -> bytes:
        """Call report generator service."""

        """
        Key remapping (orchestrator -> report_generator YAML paths):
          round1["mcgm"]       -> mcgm_property
          round1["dp_remarks"] -> dp_report
          round2["aviation_height"] -> height
          round2["ready_reckoner"]  -> ready_reckoner  (same name)
          round1["site_analysis"]   -> site_analysis   (same name)
        """
        mcgm = round1.get("mcgm", {}) or {}
        dp = round1.get("dp_remarks", {}) or {}

        # Merge GIS metrics from MCGM into DP remarks for report generator.
        # Only merge non-zero values for road_width and abutting_length — 0 means
        # "GIS layer had no feature" which is not the same as "road doesn't exist",
        # so we let the YAML fallback (18.3m / 100m) apply instead.
        if mcgm.get("setback_area_m2") is not None:
            dp["setback_area_sqm"] = dp.get("setback_area_sqm") or mcgm.get("setback_area_m2")
        if mcgm.get("max_road_width_m"):  # only if > 0
            dp["road_width_m"] = dp.get("road_width_m") or mcgm.get("max_road_width_m")
        if mcgm.get("reservation_area_m2") is not None:
            dp["reservation_area_sqm"] = dp.get("reservation_area_sqm") or mcgm.get(
                "reservation_area_m2"
            )

        if mcgm.get("zone_code"):
            dp["zone_code"] = dp.get("zone_code") or mcgm["zone_code"]
        # Only copy non-zero abutting_length — 0 means not found in GIS (let YAML fallback apply)
        if mcgm.get("abutting_length_m"):
            dp["abutting_length_m"] = mcgm.get("abutting_length_m")
        dp["nalla_present"] = mcgm.get("nalla_present")
        dp["industrial_present"] = mcgm.get("industrial_present")
        dp["roads_touching"] = mcgm.get("roads_touching")
        # 0 from MCGM means "not found in GIS layer", not "0 entrances" — let YAML fallback=2 apply
        dp["carriageway_entrances"] = mcgm.get("carriageway_entrances") or None

        pr_card = round1.get("pr_card", {}) or {}
        site = round1.get("site_analysis", {}) or {}
        aviation = round2.get("aviation_height", {}) or {}
        rr = round2.get("ready_reckoner", {}) or {}
        ocr = round1.get("ocr", {}) or {}
        tenements_ocr = round1.get("tenements_ocr", {}) or {}

        # MCGM building_data: scraped from property popup (floors, usage, etc.)
        building_data = mcgm.get("building_data") or {}
        pr_card_extracted = pr_card.get("extracted_data") or {}

        # ── Scalar fields: best-source priority ──────────────────────────────
        society_name = (
            pr_card.get("society_name")
            or mcgm.get("society_name")
            or req.get("society_name")
            or req.get("address", "Unknown Society")
        )
        # Plot area: PR Card > MCGM area_sqm > ArcGIS area > req override
        plot_area_sqm = (
            pr_card.get("area_sqm")
            or pr_card_extracted.get("area_sqm")
            or mcgm.get("area_sqm")
            or mcgm.get("plot_area_sqm")
            or req.get("plot_area_sqm")
        )
        road_width_m = dp.get("road_width_m") or site.get("road_width_m") or req.get("road_width_m")
        # num_flats: tenements sheet OCR (highest) > req (manual entry) > PR Card > MCGM > OCR old plan
        num_flats = int(
            tenements_ocr.get("num_flats")
            or req.get("num_flats")
            or pr_card.get("num_flats")
            or building_data.get("num_flats")
            or mcgm.get("num_flats")
            or ocr.get("num_flats")
            or 0
        )
        num_commercial = int(
            tenements_ocr.get("num_commercial")
            or req.get("num_commercial")
            or pr_card.get("num_commercial")
            or building_data.get("num_commercial")
            or mcgm.get("num_commercial")
            or ocr.get("num_commercial")
            or 0
        )
        # OCR-sourced carpet areas (sqft): OCR service > pr_card > req
        residential_area_sqft = (
            ocr.get("existing_residential_carpet_sqft")
            or pr_card.get("residential_area_sqft")
            or pr_card.get("existing_residential_carpet_sqft")
            or req.get("residential_area_sqft")
            or req.get("existing_residential_carpet_sqft")
        )
        commercial_area_sqft = (
            ocr.get("existing_commercial_carpet_sqft")
            or pr_card.get("commercial_area_sqft")
            or pr_card.get("existing_commercial_carpet_sqft")
            or req.get("commercial_area_sqft")
            or req.get("existing_commercial_carpet_sqft")
        )

        # society_age and existing_bua: OCR service > building_data > req
        society_age = (
            ocr.get("society_age") or building_data.get("society_age") or req.get("society_age")
        )
        existing_bua_sqft = (
            ocr.get("existing_total_bua_sqft")
            or building_data.get("existing_bua_sqft")
            or req.get("existing_bua_sqft")
            # Fallback to carpet sum * 1.2 if missing
            or ((float(residential_area_sqft or 0) + float(commercial_area_sqft or 0)) * 1.2)
        )
        # Old setback from old plan OCR (N20)
        old_setback_sqm = ocr.get("setback_area_sqm") or req.get("old_setback_sqm")

        # ── Financial rates: RR service > manual override ──────────────────────
        # RR service returns data from JSONL - no fallback needed
        caller_financial = req.get("financial") or {}
        financial = {
            "sale_rate_residential": rr.get("sale_rate_residential")
            or caller_financial.get("sale_rate_residential"),
            "sale_rate_commercial_gf": rr.get("sale_rate_commercial_gf")
            or caller_financial.get("sale_rate_commercial_gf"),
            "sale_rate_commercial_1f": caller_financial.get("sale_rate_commercial_1f") or 60000,
            "sale_rate_commercial_2f": caller_financial.get("sale_rate_commercial_2f") or 0,
            "sale_rate_commercial_other": caller_financial.get("sale_rate_commercial_other") or 0,
            "parking_price_per_unit": rr.get("parking_price_per_unit")
            or caller_financial.get("parking_price_per_unit"),
            # Merge any other caller overrides
            **{
                k: v
                for k, v in caller_financial.items()
                if k
                not in (
                    "sale_rate_residential",
                    "sale_rate_commercial_gf",
                    "parking_price_per_unit",
                )
            },
        }

        # Enrich mcgm_property blob with village and resolved identifiers
        mcgm_enriched = {
            **mcgm,
            "cts_no": mcgm.get("cts_no") or req.get("cts_no"),
            "fp_no": mcgm.get("fp_no") or req.get("fp_no"),
            "tps_name": mcgm.get("tps_name") or req.get("tps_name"),
            "village": mcgm.get("village") or req.get("village", ""),
            "area_sqm": plot_area_sqm or mcgm.get("area_sqm"),
        }

        # Derive noc_civil_aviation from aviation height result if not already in manual_inputs
        manual_inputs = dict(req.get("manual_inputs") or {})

        # Propagate old_setback_sqm from OCR into manual_inputs for N20
        if old_setback_sqm and "old_setback_sqm" not in manual_inputs:
            manual_inputs["old_setback_sqm"] = old_setback_sqm

        # Propagate top-level user fields into manual_inputs
        for _field in (
            "bankGuranteeCommercial",
            "bankGuranteeResidential",
            "costAcquisition79a",
            "salableResidentialRatePerSqFt",
            "carsToSellRatePerCar",
            "saleAreaBreakup",
            "plot_area_sqm",
            "zone_code",
            "fsi",
        ):
            if _field not in manual_inputs and req.get(_field) is not None:
                manual_inputs[_field] = req[_field]

        if "noc_civil_aviation" not in manual_inputs:
            # aviation may be {"status":..., "data":{...}} — unwrap if needed
            _avi = (
                aviation.get("data", aviation)
                if isinstance(aviation.get("data"), dict)
                else aviation
            )
            restriction = (_avi.get("restriction_reason") or "").lower()
            noc_required = bool(_avi.get("max_height_m")) or bool(
                restriction and restriction != "no restriction"
            )
            manual_inputs["noc_civil_aviation"] = 1 if noc_required else 0

        # M1: populate CTS/FP label from user-supplied identifier
        if "cts_fp_no_label" not in manual_inputs:
            cts_val = req.get("cts_no") or req.get("fp_no") or ""
            manual_inputs["cts_fp_no_label"] = f"Cts No. /FP No.:- {cts_val}".strip()

        # Nalla Note logic (B327)
        if not manual_inputs.get("nalla_note"):
            manual_inputs["nalla_note"] = 55000 if dp.get("nalla_present") else 0

        # J45: saleCommercialMunBuaSqFt is a PERCENTAGE
        commercial_bua_pct = manual_inputs.get("commercial_bua_sqft")
        if commercial_bua_pct is not None:
            # Pass as raw percentage for the formula to use
            manual_inputs["commercial_percentage"] = commercial_bua_pct
            # Remove the confusing sqft key if it was just the pct
            if manual_inputs.get("commercial_bua_sqft") == commercial_bua_pct:
                del manual_inputs["commercial_bua_sqft"]

        payload = {
            # Meta
            "society_name": society_name,
            "scheme": req.get("scheme", "33(7)(B)"),
            "redevelopment_type": req.get("redevelopment_type", "CLUBBING"),
            "ward": req.get("ward"),
            # Scalar fields (top-level, read directly by cell_mapper)
            "plot_area_sqm": plot_area_sqm,
            "road_width_m": road_width_m,
            "num_flats": num_flats,
            "num_commercial": num_commercial,
            "society_age": society_age,
            "existing_bua_sqft": existing_bua_sqft,
            # OCR-derived carpet areas and setback
            "existing_residential_carpet_sqft": residential_area_sqft,
            "existing_commercial_carpet_sqft": commercial_area_sqft,
            "old_setback_sqm": old_setback_sqm,
            # Microservice blobs — key names must match YAML `from:` paths
            "mcgm_property": mcgm_enriched,
            "dp_report": dp,
            "site_analysis": site,
            "height": aviation,
            "ready_reckoner": rr,
            # Financial rates (RR-derived + caller overrides)
            "financial": financial,
            "manual_inputs": manual_inputs,
            "premium": {},
            "zone_regulations": {},
            "fsi": {},
            "bua": {},
        }

        logger.info(
            f"[REPORT] Generating 33(7)(B) report for job {job_id} with plot_area={plot_area_sqm}, zone={manual_inputs.get('zone_code')}, fsi={manual_inputs.get('fsi')}"
        )
        resp = await client.post(
            f"{REPORT_GENERATOR_URL}/generate/feasibility-report",
            json=payload,
        )
        resp.raise_for_status()
        return resp.content

    async def _check_expiries(self, job_id: str, req: dict):
        "Check for red cells in the generated report and send notifications if close to expiry."
        try:
            import openpyxl

            from .email import send_email

            # Get report path from DB
            try:
                async with async_session_factory() as db:
                    report = await db.get(FeasibilityReport, job_id)
                    report_path = report.report_path if report else None
            except Exception:
                report_path = None

            if not report_path or not os.path.exists(report_path):
                return

            wb = openpyxl.load_workbook(report_path)
            red_found = []

            # Scan Details sheet for red cells (FFFF0000)
            if "Details" in wb.sheetnames:
                ws = wb["Details"]
                for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
                    for cell in row:
                        if (
                            getattr(cell, "fill", None)
                            and getattr(cell.fill, "fgColor", None)
                            and cell.fill.fgColor.rgb == "FFFF0000"
                        ):
                            val = str(cell.value or "")
                            red_found.append(f"{cell.coordinate}: {val}")

            if red_found:
                soc_name = req.get("society_name") or "Unknown Society"
                subject = f"⚠️ Expiry Warning: {soc_name} Feasibility Report"

                items_html = "".join([f"<li><strong>{item}</strong></li>" for item in red_found])
                body = f"""
                <html>
                <body style="font-family: sans-serif;">
                    <h2 style="color: #c62828;">Critical Expiry Warning</h2>
                    <p>The feasibility report for <strong>{soc_name}</strong> (Job: {job_id}) contains fields marked for expiry:</p>
                    <ul style="background: #ffebee; padding: 20px; border-radius: 8px; list-style: none;">
                        {items_html}
                    </ul>
                    <p>These values need to be updated as their validity period is ending soon.</p>
                    <hr>
                    <p style="font-size: 12px; color: #666;">This is an automated notification from Dhara AI.</p>
                </body>
                </html>
                """
                # Send to admin email as configured in settings
                await send_email(settings.SMTP_FROM_EMAIL, subject, body)
                logger.info(f"[{job_id}] Sent expiry notification for {len(red_found)} red cells")

        except Exception:
            logger.exception(f"[{job_id}] Expiry check failed")


# Singleton instance
feasibility_orchestrator = FeasibilityOrchestrator()
