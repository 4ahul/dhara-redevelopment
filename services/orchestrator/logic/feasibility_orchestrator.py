"""
Feasibility Analysis Orchestrator Service
Orchestrates all microservices for full feasibility analysis.
"""

import asyncio
import logging
import os
from pathlib import Path
from uuid import uuid4

from fastapi import BackgroundTasks

from dhara_shared.dhara_common.http import AsyncHTTPClient
from services.orchestrator.core.config import settings
from services.orchestrator.logic.cloudinary import upload_content
from services.orchestrator.db import async_session_factory
from services.orchestrator.models import FeasibilityReport, ReportStatus
from sqlalchemy import select

logger = logging.getLogger(__name__)

# In-memory store: job_id -> report file path (survives request lifecycle)
_REPORT_STORE: dict[str, str] = {}

# Where to write generated reports (inside the orchestrator's working dir)
_REPORTS_DIR = Path(os.getenv("REPORTS_DIR", "./generated_reports"))
_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

MCGM_URL = settings.MCGM_PROPERTY_URL
SITE_ANALYSIS_URL = settings.SITE_ANALYSIS_URL
DP_REMARKS_URL = settings.DP_REPORT_URL
PR_CARD_URL = settings.PR_CARD_URL
AVIATION_HEIGHT_URL = settings.HEIGHT_URL
READY_RECKONER_URL = settings.READY_RECKONER_URL
REPORT_GENERATOR_URL = settings.REPORT_URL
OCR_URL = settings.REPORT_URL  # OCR endpoint lives on the report_generator


class FeasibilityOrchestrator:
    """Orchestrates all microservices for feasibility analysis."""

    async def analyze(self, req: dict, background_tasks: BackgroundTasks | None = None, user_id: str | None = None, report_id: str | None = None) -> dict:
        """
        Main orchestration method. Now supports persistence via user_id and report_id.
        """
        job_id = report_id or str(uuid4())
        logger.info(f"[{job_id}] Starting feasibility analysis")

        # Step 0: Ensure a DB record exists if we have a user_id
        if user_id and not report_id:
            try:
                async with async_session_factory() as db:
                    new_report = FeasibilityReport(
                        id=job_id,
                        user_id=user_id,
                        society_id=req.get("society_id"),
                        title=req.get("title", f"Analysis: {req.get('society_name', 'Unnamed')}"),
                        status=ReportStatus.PROCESSING,
                        input_data=req
                    )
                    db.add(new_report)
                    await db.commit()
            except Exception as e:
                logger.warning(f"[{job_id}] Could not create initial report record: {e}")

        # Step 0: Resolve CTS/FP
        cts_no = req.get("cts_no")
        fp_no = req.get("fp_no")
        if cts_no or fp_no:
            from services.orchestrator.logic.cts_fp_resolver import get_resolver
            resolver = get_resolver()
            res = await resolver.resolve(
                cts_no=cts_no,
                fp_no=fp_no,
                village=req.get("village"),
                ward=req.get("ward"),
                address=req.get("address")
            )
            if res:
                if res.cts_no:
                    req["cts_no"] = res.cts_no
                if res.fp_no:
                    req["fp_no"] = res.fp_no
                if res.tps_name:
                    req["tps_name"] = res.tps_name
                # ArcGIS location data is authoritative
                if res.extra:
                    ext = res.extra
                    if not req.get("ward"):
                        req["ward"] = ext.get("ward")
                    if not req.get("village"):
                        req["village"] = ext.get("village")
                    if not req.get("taluka"):
                        req["taluka"] = ext.get("taluka")
                    if not req.get("district"):
                        req["district"] = ext.get("district")
                    if not req.get("plot_area_sqm") and ext.get("area_sqm"):
                        req["plot_area_sqm"] = float(ext["area_sqm"])
                logger.info(f"[{job_id}] CTS/FP resolved: {req.get('cts_no')} / {req.get('fp_no')}")

        # Round 1: parallel calls to 4 services
        round1_results = await self.run_round1(req, job_id=job_id)
        logger.info(f"[{job_id}] Round 1 completed")

        # Extract dependencies for Round 2
        # Try multiple sources for lat/lng (MCGM, site_analysis)
        mc_result = round1_results.get("mcgm", {})
        sa_result = round1_results.get("site_analysis", {})

        # MCGM may return centroid_lat/lng or lat/lng
        lat = mc_result.get("centroid_lat") or mc_result.get("lat") or sa_result.get("lat")
        lng = mc_result.get("centroid_lng") or mc_result.get("lng") or sa_result.get("lng")

        # DP Remarks may return zone_code or zone
        zone = round1_results.get("dp_remarks", {}).get("zone_code") or round1_results.get("dp_remarks", {}).get("zone")

        logger.info(f"[{job_id}] Round 1 extracted: lat={lat}, lng={lng}, zone={zone}")

        # Round 2: dependent services — pass req for locality context
        round2_results = await self.run_round2(lat, lng, zone, req, job_id=job_id)
        logger.info(f"[{job_id}] Round 2 completed")

        # Round 3: forward aggregated data to report generator
        report_url = None
        report_error = None
        try:
            async with AsyncHTTPClient(timeout=300.0, request_id=job_id) as client:
                report_bytes = await self.call_report_generator(
                    client, round1_results, round2_results, req
                )
            # 1. Save locally as fallback/cache
            report_filename = f"feasibility_{job_id}.xlsx"
            report_path = str(_REPORTS_DIR / report_filename)
            Path(report_path).write_bytes(report_bytes)
            _REPORT_STORE[job_id] = report_path
            
            # 2. Upload to Cloudinary for durability (Task 1)
            try:
                upload_res = await upload_content(
                    content=report_bytes,
                    filename=report_filename,
                    folder="dhara/reports",
                    resource_type="raw"
                )
                report_url = upload_res.get("secure_url")
                logger.info(f"[{job_id}] Report uploaded to Cloudinary: {report_url}")
            except Exception as ue:
                logger.warning(f"[{job_id}] Cloudinary upload failed (falling back to local): {ue}")

            logger.info(f"[{job_id}] Round 3 done — generated report for {job_id}")

            if background_tasks:
                background_tasks.add_task(self._check_expiries, job_id, req)
        except Exception as e:
            report_error = str(e)
            logger.error(f"[{job_id}] Round 3 (report generation) failed: {e}")

        # Update DB with final results if we have a job_id record
        try:
            async with async_session_factory() as db:
                report = await db.get(FeasibilityReport, job_id)
                if report:
                    report.status = ReportStatus.COMPLETED if not report_error else ReportStatus.FAILED
                    report.report_path = report_path
                    report.file_url = report_url # Cloudinary URL
                    report.output_data = {
                        "round1": round1_results,
                        "round2": round2_results,
                        "report_generated": report_path is not None,
                        "report_url": report_url,
                        "report_error": report_error
                    }
                    if report_error:
                        report.error_message = report_error
                    await db.commit()
        except Exception as e:
            logger.warning(f"[{job_id}] Could not finalize report record: {e}")

        return {
            "job_id":          job_id,
            "status":          "completed" if not report_error else "failed",
            "round1_results":  round1_results,
            "round2_results":  round2_results,
            "report_generated": report_path is not None,
            "report_url":      report_url,
            "report_error":    report_error,
        }

    async def run_round1(self, req: dict, job_id: str | None = None) -> dict:
        """Call all 4 Round 1 services in parallel."""

        async with AsyncHTTPClient(timeout=300.0, request_id=job_id) as client:
            tasks = [
                self.call_pr_card(client, req),
                self.call_mcgm(client, req),
                self.call_site_analysis(client, req),
                self.call_dp_remarks(client, req),
            ]

            results = await asyncio.gather(*tasks, return_exceptions=True)

            return {
                "pr_card": results[0] if not isinstance(results[0], Exception) else {"error": str(results[0])},
                "mcgm": results[1] if not isinstance(results[1], Exception) else {"error": str(results[1])},
                "site_analysis": results[2] if not isinstance(results[2], Exception) else {"error": str(results[2])},
                "dp_remarks": results[3] if not isinstance(results[3], Exception) else {"error": str(results[3])},
            }

    async def run_round2(
        self,
        lat: float | None,
        lng: float | None,
        zone: str | None,
        req: dict = None,
        job_id: str | None = None,
    ) -> dict:
        """Call Round 2 dependent services."""
        req = req or {}
        async with AsyncHTTPClient(timeout=300.0, request_id=job_id) as client:
            tasks = []

            # Aviation Height needs lat/lng
            if lat and lng:
                tasks.append(self.call_aviation_height(client, lat, lng))
            else:
                async def no_lat_lng():
                    return {"error": "No lat/lng available"}
                tasks.append(no_lat_lng())

            # Ready Reckoner needs zone + req context
            if zone:
                tasks.append(self.call_ready_reckoner(client, zone, req))
            else:
                async def no_zone():
                    return {"error": "No zone available"}
                tasks.append(no_zone())

            results = await asyncio.gather(*tasks, return_exceptions=True)

            return {
                "aviation_height": results[0] if not isinstance(results[0], Exception) else {"error": str(results[0])},
                "ready_reckoner": results[1] if not isinstance(results[1], Exception) else {"error": str(results[1])},
            }

    # ─── Individual service calls ─────────────────────────────────────

    async def call_pr_card(self, client: AsyncHTTPClient, req: dict) -> dict:
        """Call PR Card Scraper service."""
        try:
            resp = await client.post(
                f"{PR_CARD_URL}/scrape/sync",
                json={
                    "district": req.get("district", "Mumbai"),
                    "taluka": req.get("taluka", "Mumbai"),
                    "village": req.get("village", "").replace("-WEST", "").replace("-EAST", "").strip(),
                    "survey_no": req.get("cts_no", ""),
                    "address": req.get("address", ""),
                }
            )
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            logger.warning(f"PR Card call failed: {e}")
            return {"error": str(e)}

    async def call_mcgm(self, client: AsyncHTTPClient, req: dict) -> dict:
        """Call MCGM Property Lookup service."""
        try:
            cts_no = req.get("cts_no") or req.get("fp_no", "")
            resp = await client.post(
                f"{MCGM_URL}/lookup/sync",
                json={
                    "ward": req.get("ward", "M/E"),
                    "village": req.get("village", "").replace("-WEST", "").replace("-EAST", "").strip(),
                    "cts_no": cts_no,
                    "use_fp": req.get("use_fp_scheme", False),
                    "tps_name": req.get("tps_name"),
                }
            )
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            logger.warning(f"MCGM call failed: {e}")
            return {"error": str(e)}

    async def call_site_analysis(self, client: AsyncHTTPClient, req: dict) -> dict:
        """Call Site Analysis service."""
        try:
            resp = await client.post(
                f"{SITE_ANALYSIS_URL}/analyse",
                json={
                    "address": req.get("address", ""),
                    "ward": req.get("ward"),
                }
            )
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            logger.warning(f"Site Analysis call failed: {e}")
            return {"error": str(e)}

    async def call_dp_remarks(self, client: AsyncHTTPClient, req: dict) -> dict:
        """Call DP Remarks service.

        Strategy:
        1. Try with the CTS number first (1991 scheme).
        2. If the CTS isn't listed in the MCGM dropdown, retry using the
           FP number + TPS scheme (2034 scheme, use_fp_scheme=True).
        """
        cts_no = req.get("cts_no")
        fp_no  = req.get("fp_no", "")
        base_payload = {
            "ward":    req.get("ward", "M/E"),
            "village": req.get("village", ""),
            "lat":     req.get("lat"),
            "lng":     req.get("lng"),
        }
        try:
            # Attempt 1: CTS number
            resp = await client.post(
                f"{DP_REMARKS_URL}/fetch/sync",
                json={**base_payload, "cts_no": cts_no or fp_no, "use_fp_scheme": False},
            )
            resp.raise_for_status()
            result = resp.json()
            # If service signals CTS not found, try FP fallback
            if result.get("cts_not_found") and fp_no:
                raise ValueError("CTS not in dropdown — retrying with FP")
            return result
        except Exception as e_cts:
            logger.warning(f"DP Remarks CTS attempt failed ({e_cts}); trying FP fallback")
            try:
                resp = await client.post(
                    f"{DP_REMARKS_URL}/fetch/sync",
                    json={**base_payload, "cts_no": fp_no, "use_fp_scheme": True,
                          "tps_name": req.get("tps_name")},
                )
                resp.raise_for_status()
                return resp.json()
            except Exception as e_fp:
                logger.warning(f"DP Remarks FP fallback also failed: {e_fp}")
                return {"error": str(e_fp)}

    async def call_aviation_height(self, client: AsyncHTTPClient, lat: float, lng: float) -> dict:
        """Call Aviation Height service."""
        try:
            resp = await client.post(
                f"{AVIATION_HEIGHT_URL}/check-height",
                json={"lat": lat, "lng": lng}
            )
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            logger.warning(f"Aviation Height call failed: {e}")
            return {"error": str(e)}

    async def call_ready_reckoner(self, client: AsyncHTTPClient, zone: str, req: dict = None) -> dict:
        """Call Ready Reckoner service with proper locality + zone + plot_area inputs.
        
        Returns normalized dict with rr_open_land_sqm and sale rates extracted
        from rr_rates[] for direct use in the template.
        """
        req = req or {}
        try:
            # Derive locality from village/address for RR lookup
            village = (req.get("village") or "").lower().replace(" ", "-").replace("/", "-")
            locality = village or "bhuleshwar"  # fallback to central Mumbai

            payload = {
                "zone": str(zone),
                "locality": locality,
                "district": "mumbai",
                "taluka": "mumbai-city",
                "plot_area_sqm": req.get("plot_area_sqm") or 0.0,
                "property_type": "residential",
                "property_area_sqm": req.get("plot_area_sqm") or 0.0,
            }
            resp = await client.post(
                f"{READY_RECKONER_URL}/calculate",
                json=payload,
            )
            resp.raise_for_status()
            raw = resp.json()

            # The response is wrapped in InternalServiceResponse {status, data, error}
            data = raw.get("data") or {}
            rr_rates = data.get("rr_rates", [])

            # Normalize: extract per-category rates from rr_rates[]
            rates = {r["category"].lower(): r["value"] for r in rr_rates}
            rr_open  = rates.get("land") or rates.get("open land") or 0
            rr_res   = rates.get("residential") or rr_open
            rr_shop  = rates.get("shop") or rates.get("office") or rr_open * 1.5

            return {
                # Preserve the extracted data block
                "rr_data": data,
                # Flat keys that template YAML reads via ready_reckoner.*
                "rr_open_land_sqm":   rr_open,
                "rr_residential_sqm": rr_res,
                "rr_shop_sqm":        rr_shop,
                # Sale rates (per sqft) derived from RR land rate for the zone
                # Typical Mumbai market: ~35–45% premium over RR
                "sale_rate_residential":     round(rr_res / 10.764 * 1.4) if rr_res else 50000,
                "sale_rate_commercial_gf":   round(rr_shop / 10.764 * 1.4) if rr_shop else 75000,
                "parking_price_per_unit":    1_200_000,
            }
        except Exception as e:
            logger.warning(f"Ready Reckoner call failed: {e}")
            return {"error": str(e)}


    async def call_report_generator(
        self,
        client: AsyncHTTPClient,
        round1: dict,
        round2: dict,
        req: dict,
    ) -> bytes:
        """
        Map orchestrator results to TemplateReportRequest and call
        the report generator service.

        Key remapping (orchestrator → report_generator YAML paths):
          round1["mcgm"]       → mcgm_property
          round1["dp_remarks"] → dp_report
          round2["aviation_height"] → height
          round2["ready_reckoner"]  → ready_reckoner  (same name)
          round1["site_analysis"]   → site_analysis   (same name)
        """
        mcgm      = round1.get("mcgm", {}) or {}
        dp        = round1.get("dp_remarks", {}) or {}
        pr_card   = round1.get("pr_card", {}) or {}
        site      = round1.get("site_analysis", {}) or {}
        aviation  = round2.get("aviation_height", {}) or {}
        rr        = round2.get("ready_reckoner", {}) or {}

        # MCGM building_data: scraped from property popup (floors, usage, etc.)
        building_data = mcgm.get("building_data", {}) or {}

        # ── Scalar fields: best-source priority ──────────────────────────────
        society_name = (
            pr_card.get("society_name")
            or mcgm.get("society_name")
            or req.get("address", "Unknown Society")
        )
        # Plot area: PR Card > MCGM area_sqm > ArcGIS area > req override
        plot_area_sqm = (
            pr_card.get("area_sqm")
            or pr_card.get("extracted_data", {}).get("area_sqm")
            or mcgm.get("area_sqm")
            or mcgm.get("plot_area_sqm")
            or req.get("plot_area_sqm")
        )
        road_width_m = (
            dp.get("road_width_m")
            or site.get("road_width_m")
            or req.get("road_width_m")
        )
        # num_flats: PR Card > MCGM building_data > req
        num_flats = int(
            pr_card.get("num_flats")
            or building_data.get("num_flats")
            or mcgm.get("num_flats")
            or req.get("num_flats")
            or 0
        )
        num_commercial = int(
            pr_card.get("num_commercial")
            or building_data.get("num_commercial")
            or mcgm.get("num_commercial")
            or req.get("num_commercial")
            or 0
        )
        # OCR-sourced carpet areas (sqft)
        residential_area_sqft = (
            pr_card.get("residential_area_sqft")
            or pr_card.get("existing_residential_carpet_sqft")
            or req.get("residential_area_sqft")
            or req.get("existing_residential_carpet_sqft")
        )
        commercial_area_sqft = (
            pr_card.get("commercial_area_sqft")
            or pr_card.get("existing_commercial_carpet_sqft")
            or req.get("commercial_area_sqft")
            or req.get("existing_commercial_carpet_sqft")
        )

        # society_age and existing_bua: from building_data (OCR) or req
        society_age = (
            building_data.get("society_age")
            or req.get("society_age")
        )
        existing_bua_sqft = (
            building_data.get("existing_bua_sqft")
            or req.get("existing_bua_sqft")
            # Fallback to carpet sum * 1.2 if missing
            or ((float(residential_area_sqft or 0) + float(commercial_area_sqft or 0)) * 1.2)
        )

        # ── Financial rates: RR service > manual override > defaults ─────────
        # rr dict already has flat keys injected by _call_ready_reckoner()
        caller_financial = req.get("financial") or {}
        financial = {
            # RR-derived sale rates (per sqft)
            "sale_rate_residential":   rr.get("sale_rate_residential")   or caller_financial.get("sale_rate_residential")   or 50000,
            "sale_rate_commercial_gf": rr.get("sale_rate_commercial_gf") or caller_financial.get("sale_rate_commercial_gf") or 75000,
            "sale_rate_commercial_1f": caller_financial.get("sale_rate_commercial_1f") or 60000,
            "sale_rate_commercial_2f": caller_financial.get("sale_rate_commercial_2f") or 0,
            "sale_rate_commercial_other": caller_financial.get("sale_rate_commercial_other") or 0,
            "parking_price_per_unit":  rr.get("parking_price_per_unit")  or caller_financial.get("parking_price_per_unit")  or 1_200_000,
            # Merge any other caller overrides
            **{k: v for k, v in caller_financial.items() if k not in ("sale_rate_residential", "sale_rate_commercial_gf", "parking_price_per_unit")},
        }

        # Enrich mcgm_property blob with village and resolved identifiers
        mcgm_enriched = {
            **mcgm,
            "cts_no": mcgm.get("cts_no") or req.get("cts_no"),
            "fp_no":  mcgm.get("fp_no") or req.get("fp_no"),
            "tps_name": mcgm.get("tps_name") or req.get("tps_name"),
            "village": mcgm.get("village") or req.get("village", ""),
            "area_sqm": plot_area_sqm or mcgm.get("area_sqm"),
        }

        payload = {
            # Meta
            "society_name":  society_name,
            "scheme":        req.get("scheme", "33(7)(B)"),
            "redevelopment_type": req.get("redevelopment_type", "CLUBBING"),
            "ward":          req.get("ward"),
            # Scalar fields (top-level, read directly by cell_mapper)
            "plot_area_sqm":  plot_area_sqm,
            "road_width_m":   road_width_m,
            "num_flats":      num_flats,
            "num_commercial": num_commercial,
            "society_age":    society_age,
            "existing_bua_sqft": existing_bua_sqft,
            # OCR-derived carpet areas
            "existing_residential_carpet_sqft": residential_area_sqft,
            "existing_commercial_carpet_sqft":  commercial_area_sqft,
            # Microservice blobs — key names must match YAML `from:` paths
            "mcgm_property":  mcgm_enriched,
            "dp_report":      dp,
            "site_analysis":  site,
            "height":         aviation,
            "ready_reckoner": rr,
            # Financial rates (RR-derived + caller overrides)
            "financial":      financial,
            "manual_inputs":  req.get("manual_inputs", {}),
            "premium":        {},
            "zone_regulations": {},
            "fsi":            {},
            "bua":            {},
        }

        resp = await client.post(
            f"{REPORT_GENERATOR_URL}/generate/feasibility-report",
            json=payload,
        )
        resp.raise_for_status()
        return resp.content

    async def _check_expiries(self, job_id: str, req: dict):
        """Check for red cells in the generated report and send notifications if close to expiry."""
        try:
            import openpyxl
            from services.orchestrator.logic.email import send_email

            report_path = _REPORT_STORE.get(job_id)
            if not report_path or not os.path.exists(report_path):
                return

            wb = openpyxl.load_workbook(report_path)
            red_found = []

            # Scan Details sheet for red cells (FFFF0000)
            if "Details" in wb.sheetnames:
                ws = wb["Details"]
                for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
                    for cell in row:
                        if (getattr(cell, "fill", None) and
                            getattr(cell.fill, "fgColor", None) and
                            cell.fill.fgColor.rgb == "FFFF0000"):
                            val = str(cell.value or "")
                            red_found.append(f"{cell.coordinate}: {val}")

            if red_found:
                soc_name = req.get("society_name") or "Unknown Society"
                subject = f"⚠️ Expiry Warning: {soc_name} Feasibility Report"

                items_html = "".join([f"<li><strong>{item}</strong></li>" for item in red_found])
                body = f"""
                <html>
                <body style="font-family: sans-serif;">
                    <h2 style="color: #c62828;">Critical Expiry Warning</h2>
                    <p>The feasibility report for <strong>{soc_name}</strong> (Job: {job_id}) contains fields marked for expiry:</p>
                    <ul style="background: #ffebee; padding: 20px; border-radius: 8px; list-style: none;">
                        {items_html}
                    </ul>
                    <p>These values need to be updated as their validity period is ending soon.</p>
                    <hr>
                    <p style="font-size: 12px; color: #666;">This is an automated notification from Dhara AI.</p>
                </body>
                </html>
                """
                # Send to admin email as configured in settings
                await send_email(settings.SMTP_FROM_EMAIL, subject, body)
                logger.info(f"[{job_id}] Sent expiry notification for {len(red_found)} red cells")

        except Exception as e:
            logger.error(f"[{job_id}] Expiry check failed: {e}")


# Singleton instance
feasibility_orchestrator = FeasibilityOrchestrator()




