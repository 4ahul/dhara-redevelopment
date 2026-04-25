"""
Verify coverage between template yellow fields and cell_mapper mappings.

Run:
  python -m services.report_generator.verify_yellow_mapping [scheme] [redevelopment_type]

Defaults: scheme=33(7)(B), redevelopment_type=CLUBBING
"""

import os
import sys
from typing import Tuple

BASE_DIR = os.path.dirname(__file__)
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from services.report_generator.logic.cell_mapper import cell_mapper
import openpyxl


FORCED_TEMPLATE_NAME = "FINAL TEMPLATE _ 33 (7)(B) .xlsx"


def _is_yellow_cell(cell) -> bool:
    if not getattr(cell, "fill", None) or not getattr(cell.fill, "fgColor", None):
        return False
    color = getattr(cell.fill.fgColor, "rgb", None)
    if not color:
        return False
    c = str(color).upper()
    return ("FFFF" in c) or (c in {"FFFFFF00", "FFFF99", "FFFFCC"})


def _label_for(ws, row: int, col: int) -> str:
    # Check nearby left columns for textual label
    for check_col in (1, 2, 3):
        c = ws.cell(row=row, column=check_col)
        if c.value and isinstance(c.value, str) and len(c.value.strip()) > 2:
            return str(c.value).strip()[:60]
    # Check above
    if row > 1:
        above = ws.cell(row=row - 1, column=col)
        if above.value:
            return str(above.value).strip()[:60]
    # Fallback to coordinate
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(col)}{row}"


def main(scheme: str = "33(7)(B)", redevelopment_type: str = "CLUBBING", mode: str = "summary") -> int:
    # Load yellow cells directly from the forced template
    template_path = os.path.join(BASE_DIR, "templates", FORCED_TEMPLATE_NAME)
    wb = openpyxl.load_workbook(template_path, data_only=False)

    # Only consider the first ~8 sheets or up to P&L if present, to mirror service logic
    try:
        pl_idx = wb.sheetnames.index("Profit & Loss Statement")
        sheetnames = wb.sheetnames[: pl_idx + 1]
    except ValueError:
        sheetnames = wb.sheetnames[:8]

    yellow = {}
    for sheet in sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if _is_yellow_cell(cell):
                    key = f"{sheet}!{cell.coordinate}"
                    yellow[key] = {
                        "sheet": sheet,
                        "cell": cell.coordinate,
                        "label": _label_for(ws, cell.row, cell.column),
                        "current_value": cell.value,
                    }

    # Build mapping set from cell_mapper for the requested scheme
    maps = cell_mapper.get_mappings_for_scheme(scheme)
    mapped = {f"{m.sheet}!{m.cell}": m for m in maps if not m.is_formula}
    # Group mapping data paths by root key to validate orchestrator payload shape
    def root_of(path: str):
        return path.split(".")[0] if path else ""
    roots = {}
    for m in maps:
        r = root_of(m.data_path)
        roots[r] = roots.get(r, 0) + 1

    missing = sorted(set(yellow.keys()) - set(mapped.keys()))
    extra = sorted(set(mapped.keys()) - set(yellow.keys()))

    print(f"Template: {template_path}")
    print(f"Scheme key: {scheme}")
    print(f"Yellow cells: {len(yellow)}")
    print(f"Mapped cells: {len(mapped)}")
    print(f"Covered (mapped ∩ yellow): {len(set(yellow) & set(mapped))}")

    if mode == "list":
        # Print all yellow cells grouped by key sheets for inspection
        wanted = [
            "Details",
            "SUMMARY 1",
            "Profit & Loss Statement",
            "Construction Cost",
            "MCGM PAYMENTS",
        ]
        for sh in wanted:
            print(f"\n-- Yellow in {sh} --")
            cnt = 0
            for key in sorted(k for k in yellow.keys() if k.startswith(sh+"!")):
                f = yellow[key]
                print(f"  {key}: label='{f['label']}' current='{f['current_value']}'")
                cnt += 1
                if cnt >= 200:
                    print("  ... truncated ...")
                    break
    else:
        print(f"Missing mappings (yellow not mapped): {len(missing)}")
        for i, c in enumerate(missing[:100], 1):
            f = yellow[c]
            print(f"  M{i:02d}: {c}  label='{f['label']}'  current='{f['current_value']}'")
        if len(missing) > 100:
            print(f"  ... and {len(missing)-100} more")

        print(f"Extra mappings (not yellow in template): {len(extra)}")
        for i, c in enumerate(extra[:50], 1):
            m = mapped[c]
            print(f"  E{i:02d}: {c}  data_path='{m.data_path}' transform='{m.transform}' default='{m.default}'")
        if len(extra) > 50:
            print(f"  ... and {len(extra)-50} more")

        print("\nMapping sources by root key:")
        for k in sorted(roots.keys()):
            print(f"  {k or '(empty)'}: {roots[k]}")

    return 0


if __name__ == "__main__":
    scheme = sys.argv[1] if len(sys.argv) > 1 else "33(7)(B)"
    rd = sys.argv[2] if len(sys.argv) > 2 else "CLUBBING"
    mode = sys.argv[3] if len(sys.argv) > 3 else "summary"
    raise SystemExit(main(scheme, rd, mode))


