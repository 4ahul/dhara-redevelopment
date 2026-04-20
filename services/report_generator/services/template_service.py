"""
Template service for loading Excel feasibility templates
and managing dynamic input fields (yellow cells).
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import os
import sys

service_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(service_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
if service_dir not in sys.path:
    sys.path.insert(0, service_dir)

from core.config import settings, resolve_scheme_key
from services.cell_mapper import cell_mapper

# Testing override: all report generation must use this single template file.
# Remove to restore per-scheme template selection from SCHEME_TEMPLATE_MAP.
_FORCED_TEMPLATE_NAME: Optional[str] = (
    "FINAL TEMPLATE _ 33 (7)(B) .xlsx"
)


@dataclass
class TemplateField:
    """Represents a yellow input cell in the template."""

    sheet: str
    cell: str
    row: int
    col: int
    current_value: any
    label: str = ""


class TemplateService:
    def __init__(self):
        self._cache: Dict[str, openpyxl.Workbook] = {}
        self._field_cache: Dict[str, List[TemplateField]] = {}
        # Cache of label indexes per (scheme_key)
        self._label_index_cache: Dict[str, Dict[str, any]] = {}

    def get_template_for_scheme(
        self, scheme: str, redevelopment_type: str = "CLUBBING"
    ) -> Path:
        """Get template file path for given scheme + redevelopment type.

        Args:
            scheme: DCPR regulation key (e.g. "30(A)", "33(20)(B)")
            redevelopment_type: "CLUBBING" (default) or "INSITU"
        """
        if _FORCED_TEMPLATE_NAME is not None:
            forced_path = settings.TEMPLATES_DIR / _FORCED_TEMPLATE_NAME
            if not forced_path.exists():
                raise FileNotFoundError(f"Forced template not found: {forced_path}")
            return forced_path

        key = resolve_scheme_key(scheme, redevelopment_type)
        template_name = settings.SCHEME_TEMPLATE_MAP[key]

        template_path = settings.TEMPLATES_DIR / template_name
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        return template_path

    def _load_workbook(self, template_path: Path) -> openpyxl.Workbook:
        """Load workbook (cached)."""
        key = str(template_path)
        if key not in self._cache:
            self._cache[key] = openpyxl.load_workbook(template_path, data_only=False)
        return self._cache[key]

    def _is_yellow_cell(self, cell) -> bool:
        """Check if a cell has yellow fill."""
        if not cell.fill or not cell.fill.fgColor:
            return False
        color = cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, "rgb") else None
        if not color:
            return False
        color_str = str(color).upper()
        return "FFFF" in color_str or color_str in ["FFFFFF00", "FFFF99", "FFFFCC"]

    def _get_cell_label(self, ws, row: int, col: int) -> str:
        """Try to get a label for the cell from nearby cells."""
        col_letter = openpyxl.utils.get_column_letter(col)

        # Check if there's a label in column A or B for this row
        for check_col in [1, 2, 3]:
            cell = ws.cell(row=row, column=check_col)
            if cell.value and isinstance(cell.value, str) and len(str(cell.value)) > 2:
                return str(cell.value)[:50]

        # Check above cell
        if row > 1:
            above = ws.cell(row=row - 1, column=col)
            if above.value:
                return str(above.value)[:50]

        return f"{col_letter}{row}"

    def get_yellow_fields(
        self, scheme: str, redevelopment_type: str = "CLUBBING"
    ) -> List[TemplateField]:
        """Get all yellow (input) cells from template for given scheme."""
        key = resolve_scheme_key(scheme, redevelopment_type)
        if key in self._field_cache:
            return self._field_cache[key]

        template_path = self.get_template_for_scheme(scheme, redevelopment_type)
        wb = self._load_workbook(template_path)

        fields = []

        # Get sheets up to and including Profit & Loss Statement
        try:
            pl_idx = wb.sheetnames.index("Profit & Loss Statement")
            sheets_to_process = wb.sheetnames[: pl_idx + 1]
        except ValueError:
            sheets_to_process = wb.sheetnames[:8]  # Default first 8 sheets

        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if self._is_yellow_cell(cell):
                        label = self._get_cell_label(ws, cell.row, cell.column)
                        fields.append(
                            TemplateField(
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                row=cell.row,
                                col=cell.column,
                                current_value=cell.value,
                                label=label,
                            )
                        )

        self._field_cache[key] = fields
        return fields

    @staticmethod
    def _normalize_label(text: str) -> str:
        s = (text or "").strip().lower()
        # collapse spaces and remove trivial punctuation
        for ch in ["\n", "\t", ":", ";", ",", "|", "(", ")"]:
            s = s.replace(ch, " ")
        s = " ".join(s.split())
        return s

    def _build_label_index(
        self, scheme: str, redevelopment_type: str = "CLUBBING"
    ):
        """Build an index to resolve manual label-based inputs to cells.

        Returns a dict with:
          - by_sheet_label: {"<Sheet>|<norm_label>": "<Cell>"}
          - by_label: {"<norm_label>": [(sheet, cell), ...]}
        """
        key = resolve_scheme_key(scheme, redevelopment_type)
        if key in self._label_index_cache:
            return self._label_index_cache[key]

        fields = self.get_yellow_fields(scheme, redevelopment_type)
        by_sheet_label: Dict[str, str] = {}
        by_label: Dict[str, List[Tuple[str, str]]] = {}
        for f in fields:
            norm = self._normalize_label(f.label or f.cell)
            k = f"{f.sheet}|{norm}"
            by_sheet_label[k] = f.cell
            by_label.setdefault(norm, []).append((f.sheet, f.cell))

        out = {"by_sheet_label": by_sheet_label, "by_label": by_label}
        self._label_index_cache[key] = out
        return out

    def get_template_sheets(
        self, scheme: str, redevelopment_type: str = "CLUBBING"
    ) -> List[str]:
        """Get list of sheets to copy (up to P&L)."""
        template_path = self.get_template_for_scheme(scheme, redevelopment_type)
        wb = self._load_workbook(template_path)

        try:
            pl_idx = wb.sheetnames.index("Profit & Loss Statement")
            return wb.sheetnames[: pl_idx + 1]
        except ValueError:
            return wb.sheetnames[:8]

    def apply_values(
        self, scheme: str, values: Dict[str, any], redevelopment_type: str = "CLUBBING"
    ) -> bytes:
        """Apply user values to template and return as bytes."""
        template_path = self.get_template_for_scheme(scheme, redevelopment_type)
        wb = openpyxl.load_workbook(template_path, data_only=False)

        # ── Formula Integrity Fix ───────────────────────────────────────────
        # We no longer delete extra sheets because templates often have 
        # hidden calculation sheets or Named Ranges that formulas depend on.
        # ────────────────────────────────────────────────────────────────────

        # Apply values to yellow cells using "Sheet!Cell" composite keys
        for field in self.get_yellow_fields(scheme, redevelopment_type):
            key = f"{field.sheet}!{field.cell}"
            if key in values:
                ws = wb[field.sheet]
                ws[field.cell] = values[key]

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def zero_yellow_fields_in_place(
        self, scheme: str, redevelopment_type: str = "CLUBBING"
    ) -> int:
        """Set all yellow input cells in the template to 0 and save in place.

        Returns:
            int: Count of cells updated to 0.
        """
        template_path = self.get_template_for_scheme(scheme, redevelopment_type)
        # Load a fresh workbook instance for writing (avoid cached object side effects)
        wb = openpyxl.load_workbook(template_path, data_only=False)

        updated = 0
        fields = self.get_yellow_fields(scheme, redevelopment_type)
        for f in fields:
            ws = wb[f.sheet]
            cell = ws[f.cell]
            val = cell.value
            # Skip formulas to preserve spreadsheet logic
            if isinstance(val, str) and val.startswith("="):
                continue
            # Write 0 as requested
            cell.value = 0
            updated += 1
            # Keep field cache roughly in sync for current_value if present
            f.current_value = 0

        wb.save(template_path)

        # Refresh the workbook cache for subsequent reads
        self._cache[str(template_path)] = wb

        return updated

    def generate_full_report(
        self,
        scheme: str,
        all_data: Dict[str, any],
        output_path: Optional[str] = None,
        redevelopment_type: str = "CLUBBING",
    ) -> Tuple[bytes, str]:
        """Generate full feasibility report using template.

        Args:
            scheme: DCPR scheme (e.g., "33(20)(B)", "33(7)(B)")
            all_data: Dict containing all microservice data
            output_path: Optional path to save Excel file
            redevelopment_type: "CLUBBING" (default) or "INSITU"

        Returns:
            Tuple of (Excel bytes, file path or temp path)
        """
        # Resolve internal key for cell_mapper (e.g. "30(A)_INSITU")
        key = resolve_scheme_key(scheme, redevelopment_type)

        # Step 1: Map microservice data to yellow cells
        cell_values = cell_mapper.map_data_to_cells(key, all_data)

        # Step 2: Merge with manual cell overrides (if any use "Sheet!Cell" or bare "Cell" keys)
        manual_inputs = all_data.get("manual_inputs", {})
        if manual_inputs:
            # Build label index once for this scheme+type
            li = self._build_label_index(scheme, redevelopment_type)
            by_sheet_label = li["by_sheet_label"]
            by_label = li["by_label"]

            for mk, mv in manual_inputs.items():
                if not isinstance(mk, str):
                    continue
                key_str = mk.strip()

                # 1) Direct cell override: "Sheet!Cell"
                if "!" in key_str:
                    cell_values[key_str] = mv
                    continue

                # 2) Sheet + label: "Sheet: Label" or "Sheet|Label"
                if ":" in key_str or "|" in key_str:
                    sep = ":" if ":" in key_str else "|"
                    sheet_hint, label_hint = [p.strip() for p in key_str.split(sep, 1)]
                    norm_label = self._normalize_label(label_hint)
                    sheet_key = f"{sheet_hint}|{norm_label}"
                    if sheet_key in by_sheet_label:
                        cell = by_sheet_label[sheet_key]
                        cell_values[f"{sheet_hint}!{cell}"] = mv
                        continue

                # 3) Label only (unique across processed sheets)
                norm_label = self._normalize_label(key_str)
                matches = by_label.get(norm_label, [])
                if len(matches) == 1:
                    s, c = matches[0]
                    cell_values[f"{s}!{c}"] = mv
                    continue

                # Else: ambiguous or no match — ignore silently

        # Step 3: Apply values to template
        excel_bytes = self.apply_values(scheme, cell_values, redevelopment_type)

        # Step 4: Save to file if path provided
        file_path = None
        if output_path:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "wb") as f:
                f.write(excel_bytes)
            file_path = output_path
        else:
            # Generate temp path
            from core.config import OUTPUT_DIR

            safe_name = all_data.get("society_name", "report").replace(" ", "_")
            output_path = str(
                OUTPUT_DIR
                / f"feasibility_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "wb") as f:
                f.write(excel_bytes)
            file_path = output_path

        return excel_bytes, file_path


template_service = TemplateService()
