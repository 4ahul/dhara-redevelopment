"""
Feasibility Report Generator — Step 8
Generates a professional Excel (.xlsx) feasibility report
mirroring the Globera Engineering format from the sample PDF.

Uses openpyxl for rich formatting: merged cells, borders, colors,
number formats, multiple sheets.
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Color palette (Dhara AI Branding: Navy + Gold + White) ──────────────────
NAVY = "0C2C65"
GOLD = "C5A059"
WHITE = "FFFFFF"
LIGHT = "F2F2F2"
MID = "E1E1E1"
GREEN = "1A6B3A"
RED_HEX = "8B1A1A"
GREY = "F5F5F5"


def _font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _hdr(ws, row, col, text, width_cols=1, color=NAVY, font_color=WHITE, size=11):
    """Write a header cell (merged if width_cols > 1)."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font(bold=True, size=size, color=font_color)
    cell.fill = _fill(color)
    cell.alignment = _align("center")
    cell.border = _border()
    if width_cols > 1:
        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + width_cols - 1,
        )
    return cell


def _data(
    ws,
    row,
    col,
    value,
    bold=False,
    fill=None,
    fmt=None,
    align="left",
    color="000000",
    wrap=False,
):
    """Write a data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=color)
    cell.alignment = _align(align, wrap=wrap)
    cell.border = _border()
    if fill:
        cell.fill = _fill(fill)
    if fmt:
        cell.number_format = fmt
    return cell


def _num(ws, row, col, value, bold=False, fill=None, fmt="#,##0.00"):
    return _data(ws, row, col, value, bold=bold, fill=fill, fmt=fmt, align="right")


# ── Sheet 1: Cover / Summary ──────────────────────────────────────────────
def build_cover(ws, data: dict):
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20

    # Title
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 30

    ws.merge_cells("A2:C2")
    t = ws["A2"]
    t.value = "FEASIBILITY STUDY REPORT"
    t.font = Font(name="Calibri", bold=True, size=22, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    ws.merge_cells("A3:C3")
    s = ws["A3"]
    s.value = data.get("society_name", "Society Name")
    s.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    s.fill = _fill(GOLD)
    s.alignment = _align("center")

    # Key info table
    info = [
        ("Prepared by", "Dhara AI (Redevelopment Intelligence)"),
        ("Date", datetime.now().strftime("%d %B %Y")),
        ("Ref No.", data.get("ref_no", "GECPL/RD/2024-25/XXX")),
        ("Property", data.get("property_desc", "")),
        ("Location", data.get("location", "")),
        ("Ward", data.get("ward", "")),
        ("Zone", data.get("zone", "Residential (R)")),
        ("Plot Area (sqm)", data.get("plot_area_sqm", 0)),
        ("Road Width (m)", data.get("road_width_m", 0)),
        ("Existing Flats", data.get("num_flats", 0)),
        ("Existing Commercial", data.get("num_commercial", 0)),
    ]
    for i, (label, val) in enumerate(info, start=5):
        ws.row_dimensions[i].height = 22
        _hdr(ws, i, 1, label, color=LIGHT, font_color="000000")
        cell = ws.cell(row=i, column=2, value=val)
        cell.font = _font(bold=True)
        cell.alignment = _align("left")
        cell.border = _border()


# ── Sheet 2: Existing Areas (Annexure I) ─────────────────────────────────
def build_existing_areas(ws, data: dict):
    ws.title = "Annexure I — Areas"
    ws.sheet_view.showGridLines = False
    for col, w in [(1, 30), (2, 10), (3, 15), (4, 15)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    _hdr(ws, 1, 1, "ANNEXURE I : EXISTING AREA STATEMENT", 4, size=13)

    # Commercial
    r = 3
    _hdr(ws, r, 1, "Shops / Commercials", 4, color=GOLD, font_color=NAVY)
    r += 1
    for h, col in [
        ("Unit No.", 1),
        ("Nos.", 2),
        ("Area (Sqm)", 3),
        ("Total (Sqft)", 4),
    ]:
        _hdr(ws, r, col, h, color=MID, font_color="000000")
    r += 1

    commercials = data.get("commercial_units", [])
    for unit in commercials:
        _data(ws, r, 1, unit.get("label", ""))
        _num(ws, r, 2, unit.get("count", 1), fmt="#,##0")
        _num(ws, r, 3, unit.get("area_sqm", 0))
        _num(ws, r, 4, unit.get("total_sqft", 0))
        r += 1

    _hdr(ws, r, 1, "Area of Shops / Commercial", 2, color=LIGHT, font_color="000000")
    _num(ws, r, 3, sum(u.get("total_sqft", 0) for u in commercials), bold=True)
    r += 2

    # Residential
    _hdr(ws, r, 1, "Flats / Residential", 4, color=GOLD, font_color=NAVY)
    r += 1
    for h, col in [
        ("Unit No.", 1),
        ("Nos.", 2),
        ("Area (Sqm)", 3),
        ("Total (Sqft)", 4),
    ]:
        _hdr(ws, r, col, h, color=MID, font_color="000000")
    r += 1

    residentials = data.get("residential_units", [])
    for unit in residentials:
        _data(ws, r, 1, unit.get("label", ""))
        _num(ws, r, 2, unit.get("count", 1), fmt="#,##0")
        _num(ws, r, 3, unit.get("area_sqm", 0))
        _num(ws, r, 4, unit.get("total_sqft", 0))
        r += 1

    _hdr(ws, r, 1, "Area of Residential", 2, color=LIGHT, font_color="000000")
    _num(ws, r, 3, sum(u.get("total_sqft", 0) for u in residentials), bold=True)


# ── Sheet 3: FSI Calculation (Annexure II) ───────────────────────────────
def build_fsi(ws, data: dict):
    ws.title = "Annexure II — FSI"
    ws.sheet_view.showGridLines = False
    cols = [35, 12, 14, 14, 14, 14]
    for i, w in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _hdr(ws, 1, 1, "ANNEXURE II : FSI CALCULATION", 6, size=13)

    schemes = ["33(7)(B)", "33(20)(B)", "33(11)", "33(12)(B)"]
    r = 3
    _hdr(ws, r, 1, "Description", color=NAVY)
    for i, sch in enumerate(schemes, 2):
        _hdr(ws, r, i, sch)
    r += 1

    # FSI rows
    fsi_data = data.get("fsi", {})
    rows = [
        ("Plot area considered (Sqm)", "plot_area"),
        ("Road Width (m)", "road_width"),
        ("Zonal FSI", "zonal_fsi"),
        ("Additional FSI (Premium)", "add_fsi_premium"),
        ("Admissible TDR on road width", "tdr_road_width"),
        ("Admissible FSI (PTC)", "fsi_ptc"),
        ("Additional FSI 33(20)(B)", "add_fsi_2020b"),
        ("Total FSI", "total_fsi"),
        ("Fungible 35%", "fungible"),
        ("Total FSI Permissible", "total_fsi_permissible"),
    ]
    for label, key in rows:
        bold = key in ("total_fsi", "total_fsi_permissible")
        fill = LIGHT if bold else None
        _data(ws, r, 1, label, bold=bold, fill=fill)
        for col, sch in enumerate(schemes, 2):
            val = fsi_data.get(sch, {}).get(key, 0)
            if val:
                _num(ws, r, col, val, bold=bold, fill=fill)
            else:
                _data(ws, r, col, "-", fill=fill, align="center")
        r += 1

    r += 1
    _hdr(ws, r, 1, "BUILT-UP AREA PERMISSIBLE", 6, color=GOLD, font_color=NAVY, size=11)
    r += 1
    bua_rows = [
        ("Total Built-Up Area Permissible (Sqft)", "total_bua_sqft"),
        ("Total RERA Carpet Area Permissible (Sqft)", "rera_carpet_sqft"),
        ("Number of Parking spaces", "parking"),
        ("Total Construction Area (Sqft)", "total_constr_sqft"),
    ]
    for label, key in bua_rows:
        bold = key in ("total_bua_sqft", "rera_carpet_sqft")
        _data(ws, r, 1, label, bold=bold)
        for col, sch in enumerate(schemes, 2):
            val = data.get("bua", {}).get(sch, {}).get(key, 0)
            _num(ws, r, col, val, bold=bold)
        r += 1


# ── Sheet 4: Financial (Annexure III) ────────────────────────────────────
def build_financial(ws, data: dict):
    ws.title = "Annexure III — Financial"
    ws.sheet_view.showGridLines = False
    cols = [5, 40, 16, 16, 16, 16]
    for i, w in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _hdr(ws, 1, 1, "ANNEXURE III : FINANCIAL CALCULATION", 6, size=13)

    schemes = ["33(7)(B)", "33(20)(B)", "33(11)", "33(12)(B)"]
    r = 3
    _data(ws, r, 1, "Sr.")
    _hdr(ws, r, 2, "Description", color=NAVY)
    for i, s in enumerate(schemes, 3):
        _hdr(ws, r, i, f"{s}\nAmount (₹)", color=NAVY)
        ws.row_dimensions[r].height = 30
    r += 1

    fin = data.get("financial", {})
    sections = [
        (
            "1",
            "Construction Cost of the Project",
            None,
            [
                ("1a", "Total construction (based on current trend)", "const_total"),
                ("1b", "Cost of parking (60% of construction cost)", "parking_cost"),
                ("1d", "Total construction cost", "const_subtotal"),
                ("1e", "18% GST on construction cost", "gst"),
                ("1f", "Total construction cost with GST", "const_with_gst"),
            ],
        ),
        (
            "2",
            "Cost of FSI, TDR Clubbing & Premiums",
            None,
            [
                ("2a", "Additional FSI on payment of premium", "add_fsi_premium"),
                ("2b", "Fungible compensatory area — Residential", "fungible_res"),
                ("2c", "Staircase premium — Mumbai Suburbs", "staircase_prem"),
                ("2d", "Open space deficiency premium", "osd_premium"),
                ("2e", "Slum TDR", "slum_tdr"),
                ("2f", "General TDR", "general_tdr"),
                ("2g", "Total cost — FSI/TDR/Premiums", "fsi_tdr_total"),
            ],
        ),
        (
            "3",
            "Cost of MCGM Approvals",
            None,
            [
                ("3a", "Scrutiny / Amended plan fees", "scrutiny"),
                ("3e", "Development charges", "dev_charges"),
                ("3h", "Development cess", "dev_cess"),
                ("3j", "Land under construction (LUC) charges", "luc"),
                ("3k", "CFO scrutiny fees", "cfo"),
                ("3w", "Heritage approval & incidental cost", "heritage"),
                ("3y", "Incidental, miscellaneous, contingencies", "misc"),
                ("3", "Total MCGM charges", "mcgm_total"),
            ],
        ),
        (
            "4",
            "Professional fees (Architect, Consultants)",
            None,
            [
                ("4", "Professional fees @ ₹125/sqft", "prof_fees"),
            ],
        ),
        (
            "5",
            "Cost for Temporary Alternate Accommodation",
            None,
            [
                ("5a", "Commercial @ ₹300/sqft × 36 months", "temp_comm"),
                ("5d", "Residential @ ₹150/sqft × 36 months", "temp_res"),
                ("5g", "Total accommodation cost", "temp_total"),
            ],
        ),
        (
            "6",
            "Stamp Duty & Registration on Agreements",
            None,
            [
                ("6a", "Stamp duty on development agreement (6%)", "stamp_duty"),
                ("6", "Total stamp duty & registration", "stamp_total"),
            ],
        ),
        (
            "7",
            "TOTAL COST OF PROJECT",
            None,
            [
                ("7", "Total cost of project", "project_total"),
            ],
        ),
        (
            "8",
            "Hardship / Corpus Fund",
            None,
            [
                ("8", "Corpus fund (₹1000–1500/sqft)", "corpus"),
            ],
        ),
        (
            "9",
            "TOTAL COST OF REDEVELOPMENT PROJECT",
            None,
            [
                (
                    "9",
                    "Total cost of redevelopment (incl. corpus)",
                    "redevelopment_total",
                ),
            ],
        ),
    ]

    for sr, section_title, _, sub_rows in sections:
        is_total = sr in ("7", "9")
        # Section header
        _data(
            ws,
            r,
            1,
            sr,
            bold=True,
            fill=NAVY if is_total else GOLD,
            color=WHITE if is_total else NAVY,
        )
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cell = ws.cell(row=r, column=2, value=section_title)
        cell.font = _font(bold=True, color=WHITE if is_total else NAVY, size=12 if is_total else 11)
        cell.fill = _fill(NAVY if is_total else GOLD)
        cell.border = _border()
        ws.row_dimensions[r].height = 22
        r += 1

        for sub_sr, desc, key in sub_rows:
            is_sub_total = key and "total" in key
            fill = LIGHT if is_sub_total else None
            _data(ws, r, 1, sub_sr, fill=fill)
            _data(ws, r, 2, desc, bold=is_sub_total, fill=fill)
            for col, sch in enumerate(schemes, 3):
                val = fin.get(sch, {}).get(key, 0) if key else 0
                if val:
                    _num(ws, r, col, val, bold=is_sub_total, fill=fill, fmt="₹#,##0")
                else:
                    _data(ws, r, col, "-", fill=fill, align="center")
            r += 1
        r += 1


# ── Sheet 5: Additional Area & Profit (Annexure IV) ───────────────────────
def build_additional_area(ws, data: dict):
    ws.title = "Annexure IV — Sale & Profit"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([5, 45, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _hdr(ws, 1, 1, "ADDITIONAL AREA ENTITLEMENT & PROFIT SUMMARY", 6, size=13)

    schemes = ["33(7)(B)", "33(20)(B)", "33(11)", "33(12)(B)"]
    r = 3
    _data(ws, r, 1, "#")
    _hdr(ws, r, 2, "Description", color=NAVY)
    for i, s in enumerate(schemes, 3):
        _hdr(ws, r, i, s)
    r += 1

    ae = data.get("additional_entitlement", {})
    rows = [
        ("1", "Cost of project (₹ Crore)", "cost_crore"),
        ("2", "Total RERA carpet area incl. fungible (Sqft)", "rera_total_sqft"),
        ("3", "Existing carpet area (Sqft)", "existing_sqft"),
        ("4", "Income per Sqft on sale (₹)", "sale_rate"),
        ("5", "Additional RERA area % considered", "add_rera_pct"),
        ("6", "Additional area for sale (Sqft)", "add_area_sqft"),
        ("7", "RERA carpet area available for sale (Sqft)", "sale_area_sqft"),
        ("8", "Revenue from project (₹ Crore)", "revenue_crore"),
        ("9", "GST for existing members (₹ Crore)", "gst_crore"),
        ("10", "Profit (₹ Crore)", "profit_crore"),
        ("11", "Profit %", "profit_pct"),
    ]
    for sr, label, key in rows:
        is_highlight = key in ("profit_crore", "profit_pct", "revenue_crore")
        fill = MID if is_highlight else None
        _data(ws, r, 1, sr, bold=is_highlight, fill=fill)
        _data(ws, r, 2, label, bold=is_highlight, fill=fill)
        for col, sch in enumerate(schemes, 3):
            val = ae.get(sch, {}).get(key, "-")
            if isinstance(val, (int, float)):
                fmt = "0.00%" if "pct" in key else "#,##0.00"
                _num(ws, r, col, val, bold=is_highlight, fill=fill, fmt=fmt)
            else:
                _data(ws, r, col, val, bold=is_highlight, fill=fill, align="center")
        r += 1


# ── Sheet 6: Site Analysis ────────────────────────────────────────────────
def build_site_analysis(ws, data: dict):
    ws.title = "Site Analysis"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50

    _hdr(ws, 1, 1, "SITE ANALYSIS & TECHNICAL PARAMETERS", 2, size=13)
    r = 3

    site = data.get("site_analysis", {})
    height = data.get("height", {})
    # Ready Reckoner data is now merged into premium checker
    premium = data.get("premium", {})
    rr = data.get("ready_reckoner", {}) or premium.get("rr_data", {})
    zone = data.get("zone_regulations", {})
    dp = data.get("dp_report", {})

    sections = [
        (
            "Google Maps Site Analysis",
            [
                ("Latitude", site.get("lat", "")),
                ("Longitude", site.get("lng", "")),
                ("Formatted Address", site.get("formatted_address", "")),
                ("Area Type", site.get("area_type", "")),
                ("Zone Inference", site.get("zone_inference", "")),
                (
                    "Nearby Landmarks",
                    ", ".join(site.get("nearby_landmarks", []))
                    if site.get("nearby_landmarks")
                    else "",
                ),
                ("Google Maps URL", site.get("maps_url", "")),
            ],
        ),
        (
            "NOCAS Height Restrictions (AAI)",
            [
                ("Max Permissible Height (m)", height.get("max_height_m", "")),
                ("Max Floors (approx.)", height.get("max_floors", "")),
                ("AAI Zone", height.get("aai_zone", "")),
                ("Restriction Reason", height.get("restriction_reason", "")),
                ("NOCAS Reference", height.get("nocas_reference", "")),
                ("RL Datum (m)", height.get("rl_datum_m", "")),
            ],
        ),
        (
            "Ready Reckoner Rates 2024-25",
            [
                ("Ward", rr.get("ward_name", premium.get("ward", ""))),
                ("RR Open Land (₹/sqm)", rr.get("rr_open_land_sqm", "")),
                ("RR Residential Building (₹/sqm)", rr.get("rr_residential_sqm", "")),
                (
                    "RR Commercial Ground (₹/sqm)",
                    rr.get("rr_commercial_ground_sqm", ""),
                ),
                (
                    "RR Construction Cost (₹/sqm)",
                    rr.get("rr_construction_cost_sqm", ""),
                ),
                ("Source", rr.get("source", "IGR Maharashtra")),
            ],
        ),
        (
            "Property Valuation (RR Formula)",
            [
                ("Property Area (sqm)", premium.get("property_area_sqm", "")),
                ("Property Type", premium.get("property_type", "").title()),
                (
                    "Amenities Premium (%)",
                    premium.get("amenities_premium_percentage", ""),
                ),
                ("Depreciation (%)", premium.get("depreciation_percentage", "")),
                ("Total Property Value (₹)", premium.get("total_property_value", "")),
            ],
        ),
        (
            "Zone Regulations (CRZ, COD, DCPR)",
            [
                ("CRZ Zone", zone.get("crz_zone", "")),
                ("CRZ Status", zone.get("crz_status", "")),
                ("CRZ NOC Required", "Yes" if zone.get("crz_noc_required") else "No"),
                ("COD Area", "Yes" if zone.get("cod_area") else "No"),
                ("DCPR Zone", zone.get("dcpr_zone", "")),
                ("Max FSI Allowed", zone.get("max_fsi", "")),
                ("Setback Requirements", zone.get("setback_requirements", "")),
                ("Parking Requirements", zone.get("parking_requirements", "")),
                ("Special Designation", zone.get("special_designation", "")),
                ("Remarks", zone.get("remarks", "")),
            ],
        ),
        (
            "Development Plan (DP) 2034",
            [
                ("DP Ward", dp.get("ward", "")),
                ("Taluka", dp.get("taluka", "")),
                ("Zone", dp.get("zone", "")),
                ("Sub Zone", dp.get("sub_zone", "")),
                ("DP Year", dp.get("dp_year", "")),
                ("Building Permission Zone", dp.get("building_permission_zone", "")),
                ("Road Width Proposed (m)", dp.get("road_width_proposed", "")),
                (
                    "Allowed Uses",
                    ", ".join(dp.get("allowed_uses", [])) if dp.get("allowed_uses") else "",
                ),
                (
                    "Restrictions",
                    "; ".join(dp.get("restrictions", [])) if dp.get("restrictions") else "",
                ),
                ("DP Remarks", dp.get("remarks", "")),
            ],
        ),
        (
            "Government Premiums & Charges",
            [
                (
                    "Total FSI/TDR Premiums (₹)",
                    premium.get("total_fsi_tdr_premiums", ""),
                ),
                ("Total MCGM Charges (₹)", premium.get("total_mcgm_charges", "")),
                ("Grand Total (₹)", premium.get("grand_total", "")),
                ("Grand Total (₹ Crore)", premium.get("grand_total_crore", "")),
                ("Scheme", premium.get("scheme", "")),
            ],
        ),
    ]

    for section_title, items in sections:
        _hdr(ws, r, 1, section_title, 2, color=GOLD, font_color=NAVY)
        r += 1
        for label, val in items:
            _hdr(ws, r, 1, label, color=LIGHT, font_color="000000")
            cell = ws.cell(row=r, column=2, value=val)
            cell.font = _font()
            cell.alignment = _align("left", wrap=True)
            cell.border = _border()
            ws.row_dimensions[r].height = 18
            r += 1
        r += 1


# ── Sheet 7: Regulatory Sources (Annexure V) ─────────────────────────────
def build_regulatory_sources(ws, data: dict):
    ws.title = "Annexure V — Regulations"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    _hdr(ws, 1, 1, "ANNEXURE V : REGULATORY SOURCES & CITATIONS", 2, size=13)
    r = 3

    # RAG sources are expected in data['regulatory_sources']
    sources = data.get("regulatory_sources", [])
    if not sources:
        _data(
            ws,
            r,
            1,
            "No specific regulatory sources cited for this analysis.",
            bold=True,
        )
        return

    _hdr(ws, r, 1, "Clause / Source", color=NAVY)
    _hdr(ws, r, 2, "Description / Rule", color=NAVY)
    r += 1

    for item in sources:
        _data(ws, r, 1, item.get("clause", "N/A"), bold=True, align="center")
        _data(ws, r, 2, item.get("text", ""), wrap=True)
        ws.row_dimensions[r].height = 60  # Allow space for text
        r += 1


# ── Sheet 8: LLM Analysis ─────────────────────────────────────────────────
def build_llm_analysis(ws, data: dict):
    ws.title = "AI Analysis"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    _hdr(ws, 1, 1, "AI-GENERATED FEASIBILITY ANALYSIS", 1, size=13)

    ws.merge_cells("A3:A50")
    cell = ws["A3"]
    cell.value = data.get("llm_analysis", "Analysis not available.")
    cell.font = _font(size=11)
    cell.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[3].height = 400


# ── Main builder ─────────────────────────────────────────────────────────
def build_feasibility_report(data: dict, output_path: str) -> str:
    """
    Build the complete feasibility report Excel file.
    Returns the path to the generated file.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_cover(wb.create_sheet(), data)
    build_existing_areas(wb.create_sheet(), data)
    build_fsi(wb.create_sheet(), data)
    build_financial(wb.create_sheet(), data)
    build_additional_area(wb.create_sheet(), data)
    build_site_analysis(wb.create_sheet(), data)
    build_regulatory_sources(wb.create_sheet(), data)
    build_llm_analysis(wb.create_sheet(), data)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
