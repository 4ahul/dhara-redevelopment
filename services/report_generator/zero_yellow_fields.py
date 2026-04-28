"""
Standalone utility to zero out all yellow input cells in the target template.

Usage:
  python -m services.report_generator.zero_yellow_fields

This will modify the template in-place:
  services/report_generator/templates/FINAL TEMPLATE _ 33 (7)(B) .xlsx
"""

import os

import openpyxl


def _is_yellow_cell(cell) -> bool:
    if not getattr(cell, "fill", None) or not getattr(cell.fill, "fgColor", None):
        return False
    color = getattr(cell.fill.fgColor, "rgb", None)
    if not color:
        return False
    c = str(color).upper()
    return ("FFFF" in c) or (c in {"FFFFFF00", "FFFF99", "FFFFCC"})


def zero_template_yellow_fields(template_path: str) -> int:
    wb = openpyxl.load_workbook(template_path, data_only=False)
    updated = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                try:
                    if _is_yellow_cell(cell):
                        val = cell.value
                        if isinstance(val, str) and val.startswith("="):
                            # keep formulas intact
                            continue
                        cell.value = 0
                        updated += 1
                except Exception:
                    # best-effort; skip problematic cells
                    pass
    wb.save(template_path)
    return updated


def main():
    base_dir = os.path.dirname(__file__)
    template_path = os.path.join(base_dir, "templates", "FINAL TEMPLATE _ 33 (7)(B) .xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")
    count = zero_template_yellow_fields(template_path)
    print(f"Zeroed {count} yellow fields in: {template_path}")


if __name__ == "__main__":
    main()
