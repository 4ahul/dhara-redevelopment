"""One-shot template inspector.

Scans an Excel template and emits a reviewable JSON dossier describing every
yellow-filled (input) and black-filled (computed) cell plus six context
signals per cell.
"""

from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook

YELLOW_RGB = "FFFFFF00"
BLACK_RGB = "FF000000"


def _fill_rgb(cell) -> str | None:
    f = cell.fill
    if getattr(f, "patternType", None) != "solid":
        return None
    fg = f.fgColor
    rgb = getattr(fg, "rgb", None) if fg else None
    return str(rgb).upper() if rgb else None


def enumerate_fillable_cells(wb: Workbook) -> list[dict[str, Any]]:
    """Return one record per yellow- or black-filled cell across every sheet."""
    out: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                rgb = _fill_rgb(cell)
                if rgb == YELLOW_RGB:
                    kind = "yellow"
                elif rgb == BLACK_RGB:
                    kind = "black"
                else:
                    continue
                out.append(
                    {
                        "sheet": sheet_name,
                        "coord": cell.coordinate,
                        "row": cell.row,
                        "col": cell.column,
                        "kind": kind,
                        "fill_rgb": rgb,
                        "current_value": cell.value,
                    }
                )
    return out


def _text(cell) -> str | None:
    v = cell.value
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return None


def _scan_left(ws, row: int, col: int) -> str | None:
    for c in range(col - 1, 0, -1):
        t = _text(ws.cell(row=row, column=c))
        if t:
            return t
    return None


def _scan_up(ws, row: int, col: int, require_bold: bool = False) -> str | None:
    for r in range(row - 1, 0, -1):
        cell = ws.cell(row=r, column=col)
        if require_bold and not (cell.font and cell.font.bold):
            continue
        t = _text(cell)
        if t:
            return t
    return None


def _merged_master(ws, cell) -> str | None:
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            master = ws.cell(row=rng.min_row, column=rng.min_col)
            if master.coordinate == cell.coordinate:
                return None
            return _text(master)
    return None


def _neighbor_3x3(ws, row: int, col: int) -> list:
    grid = []
    for dr in (-1, 0, 1):
        line = []
        for dc in (-1, 0, 1):
            r, c = row + dr, col + dc
            if r < 1 or c < 1:
                line.append(None)
            else:
                line.append(_text(ws.cell(row=r, column=c)))
        grid.append(line)
    return grid


def extract_signals(ws, cell) -> dict[str, Any]:
    """Return the six context signals for a single cell."""
    return {
        "placeholder_text": _text(cell),
        "row_label": _scan_left(ws, cell.row, cell.column),
        "column_header": _scan_up(ws, cell.row, cell.column, require_bold=True),
        "section_header": _scan_up(ws, cell.row, 1, require_bold=True),
        "merged_master": _merged_master(ws, cell),
        "neighbor_3x3": _neighbor_3x3(ws, cell.row, cell.column),
    }


def _slug(text: str) -> str:
    t = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return t or "unknown"


def suggest_mapping(kind: str, signals: dict[str, Any]) -> dict[str, Any]:
    """Return {suggested_semantic_name, suggested_source, review_required}.

    Heuristic hints only — human review is authoritative.
    """
    ph = (signals.get("placeholder_text") or "").lower()
    row = signals.get("row_label") or ""
    sect = signals.get("section_header") or ""
    header = signals.get("column_header") or ""

    # Review required if all 5 primary signals are empty.
    primaries = [
        signals.get(k)
        for k in (
            "placeholder_text",
            "row_label",
            "section_header",
            "column_header",
            "merged_master",
        )
    ]
    review = not any(primaries)

    # Build a semantic name guess from the most specific available label.
    name_seed = row or header or sect or signals.get("merged_master") or "cell"
    semantic = _slug(name_seed)

    suggested = None
    if kind == "black":
        if "mark 1" in ph or "if yes" in ph:
            noc_slug = _slug(row) if row else "unknown"
            suggested = f"calc: noc_flag_from_dp(noc_type={noc_slug})"
        elif "15%" in ph or "annually of user input" in ph:
            suggested = "calc: bank_guarantee_15pct(based_on=<fill_in>)"
        elif "calculate no of floors" in ph or "per floor is 3mtr" in ph:
            suggested = "calc: floors_from_max_height(based_on=max_height_m)"
    elif "ready reckoner" in ph or "rr " in ph:
        suggested = f"from: ready_reckoner.{semantic}"
    elif "user input" in ph or "manual" in ph:
        suggested = f"from: manual_inputs.{semantic}"
    elif "dp remark" in ph or "dp report" in ph:
        suggested = f"from: dp_report.{semantic}"
    elif "ocr" in ph or "old plan" in ph or "pr card" in ph:
        suggested = f"from: mcgm_property.{semantic}"

    if not suggested:
        suggested = "TODO: human review"
        review = True

    return {
        "suggested_semantic_name": semantic,
        "suggested_source": suggested,
        "review_required": review,
    }


def build_dossier(template_path: str, scheme: str, out_path: str | None = None) -> dict[str, Any]:
    wb = openpyxl.load_workbook(template_path, data_only=False)
    cells_raw = enumerate_fillable_cells(wb)
    cells_out: list[dict[str, Any]] = []
    for rec in cells_raw:
        ws = wb[rec["sheet"]]
        cell = ws[rec["coord"]]
        signals = extract_signals(ws, cell)
        suggestion = suggest_mapping(kind=rec["kind"], signals=signals)
        is_formula = isinstance(rec["current_value"], str) and rec["current_value"].startswith("=")
        cells_out.append(
            {
                "cell": f"{rec['sheet']}!{rec['coord']}",
                "kind": rec["kind"],
                "fill_rgb": rec["fill_rgb"],
                "current_value": rec["current_value"],
                "is_formula": is_formula,
                "signals": signals,
                "review": suggestion,
            }
        )

    dossier = {
        "template": Path(template_path).as_posix(),
        "scheme": scheme,
        "generated_at": date.today().isoformat(),
        "cells": cells_out,
    }

    if out_path:
        Path(out_path).parent.mkdir(parents=True, exist_ok=True)
        Path(out_path).write_text(json.dumps(dossier, indent=2, default=str))
    return dossier


def main(argv: list[str] | None = None) -> int:
    import argparse

    p = argparse.ArgumentParser(description="Build dossier for a feasibility template.")
    p.add_argument("--template", required=True)
    p.add_argument("--scheme", required=True)
    p.add_argument("--out", required=True)
    args = p.parse_args(argv)

    build_dossier(args.template, args.scheme, args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
