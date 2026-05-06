import openpyxl
from openpyxl.styles import Font, PatternFill

from services.report_generator.feasibility.writer import Writer


def test_writes_non_formula_cells(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Details"
    ws["A1"].value = None
    w = Writer()
    w.stage("Details!A1", 123)
    report = w.flush(wb)
    assert ws["A1"].value == 123
    assert report["skipped_formula_cells"] == []


def test_skips_formula_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Details"
    ws["A1"].value = "=B1+1"
    w = Writer()
    w.stage("Details!A1", 999)
    report = w.flush(wb)
    assert ws["A1"].value == "=B1+1"
    assert "Details!A1" in report["skipped_formula_cells"]


def test_preserves_style():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Details"
    ws["A1"].fill = PatternFill("solid", fgColor="FFFFFF00")
    ws["A1"].font = Font(bold=True)
    w = Writer()
    w.stage("Details!A1", 42)
    w.flush(wb)
    assert ws["A1"].fill.fgColor.rgb == "FFFFFF00"
    assert ws["A1"].font.bold is True
