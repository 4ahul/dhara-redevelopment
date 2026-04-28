from services.report_generator.feasibility.dispatcher import apply_transform

# noqa: E402


def test_transform_float():
    assert apply_transform("3.14", "float") == 3.14


def test_transform_int():
    assert apply_transform("42", "int") == 42


def test_transform_str():
    assert apply_transform(42, "str") == "42"


def test_transform_bool_toggle():
    assert apply_transform(True, "bool_toggle") == 1
    assert apply_transform(False, "bool_toggle") == 0
    assert apply_transform("yes", "bool_toggle") == 1
    assert apply_transform("no", "bool_toggle") == 0


def test_transform_percent():
    assert apply_transform(50, "percent") == 0.5


def test_transform_none_passthrough():
    assert apply_transform(None, "float") is None


def test_transform_none_type_passthrough():
    assert apply_transform("abc", None) == "abc"


import pytest  # noqa: E402

from services.report_generator.feasibility.calc_registry import (  # noqa: E402
    _clear_for_tests,
    register,
)
from services.report_generator.feasibility.dispatcher import resolve_entry  # noqa: E402
from services.report_generator.feasibility.exceptions import MissingData  # noqa: E402
from services.report_generator.feasibility.mapping_loader import MappingEntry  # noqa: E402


def setup_function():
    _clear_for_tests()


def test_resolve_from_path():
    e = MappingEntry("Details!A1", "yellow", "a", from_="x.y", fallback=0)
    ctx = {"request": {"x": {"y": 7}}, "resolved": {}}
    assert resolve_entry(e, ctx) == 7


def test_resolve_sources_first_non_none():
    e = MappingEntry("Details!A1", "yellow", "a", sources=["x.y", "x.z"], fallback=0)
    ctx = {"request": {"x": {"z": 3}}, "resolved": {}}
    assert resolve_entry(e, ctx) == 3


def test_resolve_all_none_raises():
    e = MappingEntry("Details!A1", "yellow", "a", from_="nope", fallback=0)
    with pytest.raises(MissingData):
        resolve_entry(e, {"request": {}, "resolved": {}})


def test_resolve_const():
    e = MappingEntry("Details!A1", "yellow", "a", const=1.28, fallback=0)
    assert resolve_entry(e, {"request": {}, "resolved": {}}) == 1.28


def test_resolve_calc():
    @register("times2")
    def times2(ctx, x: float):
        return float(x) * 2

    e = MappingEntry("Details!A1", "black", "a", calc="times2", calc_args={"x": 5}, fallback=0)
    assert resolve_entry(e, {"request": {}, "resolved": {}}) == 10.0


def test_generate_end_to_end(tmp_path):
    """Build a tiny template + mapping, run dispatcher.generate, check xlsx."""
    import openpyxl
    from openpyxl.styles import PatternFill

    # Build template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Details"
    ws["A1"].fill = PatternFill("solid", fgColor="FFFFFF00")  # yellow
    ws["B1"].fill = PatternFill("solid", fgColor="FF000000")  # black
    tpl = tmp_path / "t.xlsx"
    wb.save(tpl)

    # Build mapping yaml
    mapping_yaml = f"""
template: {tpl.name}
scheme: "33(7)(B)"
cells:
  - cell: Details!A1
    kind: yellow
    semantic_name: road_width_m
    from: dp_report.road_width_m
    fallback: 0
    transform: float
  - cell: Details!B1
    kind: black
    semantic_name: doubled
    calc: times2_for_dispatcher
    calc_args: {{x: 10}}
    fallback: 0
    transform: float
"""
    mf_path = tmp_path / "m.yaml"
    mf_path.write_text(mapping_yaml)

    # Register the calc
    from services.report_generator.feasibility.calc_registry import _clear_for_tests, register

    _clear_for_tests()

    @register("times2_for_dispatcher")
    def times2(ctx, x):
        return x * 2

    from services.report_generator.feasibility.dispatcher import generate

    req = {"dp_report": {"road_width_m": 18.3}}
    resp = generate(
        request=req, mapping_path=str(mf_path), template_path=str(tpl), output_path=str(tpl)
    )
    assert resp.cells_written == 2
    assert resp.missing_fields == []

    # Reopen and verify
    wb2 = openpyxl.load_workbook(tpl, data_only=False)
    assert wb2["Details"]["A1"].value == 18.3
    assert wb2["Details"]["B1"].value == 20.0
