import openpyxl
from openpyxl.styles import PatternFill
from feasibility.inspector import enumerate_fillable_cells


def _make_wb(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Details"
    yellow = PatternFill("solid", fgColor="FFFFFF00")
    black = PatternFill("solid", fgColor="FF000000")
    ws["A1"].value = "label"
    ws["B1"].value = "Input"
    ws["B1"].fill = yellow
    ws["C1"].value = "Output"
    ws["C1"].fill = black
    ws["D1"].value = "=B1+1"
    path = tmp_path / "t.xlsx"
    wb.save(path)
    return path


def test_enumerate_yellow_and_black_only(tmp_path):
    path = _make_wb(tmp_path)
    wb = openpyxl.load_workbook(path, data_only=False)
    cells = enumerate_fillable_cells(wb)
    coords = {(c["sheet"], c["coord"], c["kind"]) for c in cells}
    assert ("Details", "B1", "yellow") in coords
    assert ("Details", "C1", "black") in coords
    assert not any(c["coord"] == "A1" for c in cells)
    assert not any(c["coord"] == "D1" for c in cells)


def test_gradient_fill_is_ignored(tmp_path):
    """Cells with GradientFill must not crash enumeration."""
    from openpyxl.styles import GradientFill
    from feasibility.inspector import _fill_rgb

    class _Cell:
        fill = GradientFill()

    assert _fill_rgb(_Cell()) is None


def test_context_signals_for_labelled_cell(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Details"
    ws["A1"].value = "NOC REQUIREMENTS"  # section header, bold
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws["A3"].value = "Highway NOC"       # row label
    yellow = PatternFill("solid", fgColor="FFFFFF00")
    ws["D3"].value = "DP remark report if yes mark 1 otherwise 0"
    ws["D3"].fill = yellow
    path = tmp_path / "ctx.xlsx"
    wb.save(path)

    from feasibility.inspector import extract_signals
    wb2 = openpyxl.load_workbook(path, data_only=False)
    ws2 = wb2["Details"]
    sig = extract_signals(ws2, ws2["D3"])
    assert sig["placeholder_text"] == "DP remark report if yes mark 1 otherwise 0"
    assert sig["row_label"] == "Highway NOC"
    assert sig["section_header"] == "NOC REQUIREMENTS"


def test_context_signals_for_unlabelled_cell(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Details"
    yellow = PatternFill("solid", fgColor="FFFFFF00")
    ws["D3"].value = None
    ws["D3"].fill = yellow
    path = tmp_path / "noctx.xlsx"
    wb.save(path)

    from feasibility.inspector import extract_signals
    wb2 = openpyxl.load_workbook(path, data_only=False)
    ws2 = wb2["Details"]
    sig = extract_signals(ws2, ws2["D3"])
    assert sig["placeholder_text"] is None
    assert sig["row_label"] is None
    assert sig["section_header"] is None
    assert isinstance(sig["neighbor_3x3"], list) and len(sig["neighbor_3x3"]) == 3


def test_suggest_user_input_from_placeholder():
    from feasibility.inspector import suggest_mapping
    signals = {"placeholder_text": "User input", "row_label": "Corpus Fund", "section_header": None, "column_header": None, "merged_master": None, "neighbor_3x3": []}
    s = suggest_mapping(kind="yellow", signals=signals)
    assert s["suggested_source"].startswith("from: manual_inputs.")
    assert s["review_required"] is False


def test_suggest_noc_flag_from_placeholder():
    from feasibility.inspector import suggest_mapping
    signals = {"placeholder_text": "DP remark report if yes mark 1 otherwise 0", "row_label": "Highway NOC", "section_header": "NOC REQUIREMENTS", "column_header": None, "merged_master": None, "neighbor_3x3": []}
    s = suggest_mapping(kind="black", signals=signals)
    assert s["suggested_source"].startswith("calc: noc_flag_from_dp")
    assert "highway" in s["suggested_source"].lower()


def test_review_required_when_no_signals():
    from feasibility.inspector import suggest_mapping
    signals = {"placeholder_text": None, "row_label": None, "section_header": None, "column_header": None, "merged_master": None, "neighbor_3x3": [[None]*3]*3}
    s = suggest_mapping(kind="yellow", signals=signals)
    assert s["review_required"] is True
