from feasibility.mapping_loader import MappingEntry, MappingFile


def test_mapping_entry_roundtrip():
    e = MappingEntry(
        cell="Details!R17",
        kind="yellow",
        semantic_name="road_width_m",
        from_="dp_report.road_width_m",
        fallback=18.3,
        transform="float",
    )
    assert e.cell == "Details!R17"
    assert e.sheet == "Details"
    assert e.coord == "R17"
    assert e.from_ == "dp_report.road_width_m"


def test_mapping_file_container():
    m = MappingFile(template="x.xlsx", scheme="33(7)(B)", cells=[])
    assert m.cells == []


def test_load_valid_yaml(tmp_path):
    from feasibility.mapping_loader import load_mapping
    y = """
template: t.xlsx
scheme: "33(7)(B)"
cells:
  - cell: Details!R17
    kind: yellow
    semantic_name: road_width_m
    from: dp_report.road_width_m
    fallback: 18.3
    transform: float
"""
    p = tmp_path / "m.yaml"
    p.write_text(y)
    m = load_mapping(str(p))
    assert m.scheme == "33(7)(B)"
    assert len(m.cells) == 1
    assert m.cells[0].from_ == "dp_report.road_width_m"
    assert m.cells[0].fallback == 18.3


import pytest
from feasibility.exceptions import MappingError


def _base_raw():
    return {
        "cell": "Details!A1",
        "kind": "yellow",
        "semantic_name": "name_x",
        "from": "a.b",
    }


def test_validate_exactly_one_value_source():
    from feasibility.mapping_loader import validate_entry_shape
    raw = _base_raw()
    raw["calc"] = "something"
    with pytest.raises(MappingError, match="exactly one"):
        validate_entry_shape(raw)


def test_validate_none_value_source():
    from feasibility.mapping_loader import validate_entry_shape
    raw = _base_raw()
    raw.pop("from")
    with pytest.raises(MappingError, match="exactly one"):
        validate_entry_shape(raw)


def test_validate_kind_value():
    from feasibility.mapping_loader import validate_entry_shape
    raw = _base_raw()
    raw["kind"] = "red"
    with pytest.raises(MappingError, match="kind"):
        validate_entry_shape(raw)


def test_validate_duplicate_cells(tmp_path):
    from feasibility.mapping_loader import load_mapping
    y = """
template: t.xlsx
scheme: x
cells:
  - {cell: Details!A1, kind: yellow, semantic_name: n1, from: a.b, fallback: 0}
  - {cell: Details!A1, kind: yellow, semantic_name: n2, from: a.c, fallback: 0}
"""
    p = tmp_path / "m.yaml"
    p.write_text(y)
    with pytest.raises(MappingError, match="Duplicate cell"):
        load_mapping(str(p))


def test_validate_duplicate_semantic_names(tmp_path):
    from feasibility.mapping_loader import load_mapping
    y = """
template: t.xlsx
scheme: x
cells:
  - {cell: Details!A1, kind: yellow, semantic_name: n1, from: a.b, fallback: 0}
  - {cell: Details!A2, kind: yellow, semantic_name: n1, from: a.c, fallback: 0}
"""
    p = tmp_path / "m.yaml"
    p.write_text(y)
    with pytest.raises(MappingError, match="Duplicate semantic_name"):
        load_mapping(str(p))


def test_topological_sort_yellows_first():
    from feasibility.mapping_loader import topological_sort, MappingEntry
    entries = [
        MappingEntry("Details!A2", "black", "b", calc="c1", calc_args={"based_on": "a"}, fallback=0),
        MappingEntry("Details!A1", "yellow", "a", from_="x.y", fallback=0),
    ]
    sorted_ = topological_sort(entries)
    assert sorted_[0].semantic_name == "a"
    assert sorted_[1].semantic_name == "b"


def test_topological_sort_detects_cycle():
    from feasibility.mapping_loader import topological_sort, MappingEntry
    entries = [
        MappingEntry("Details!A1", "black", "a", calc="c", calc_args={"based_on": "b"}, fallback=0),
        MappingEntry("Details!A2", "black", "b", calc="c", calc_args={"based_on": "a"}, fallback=0),
    ]
    with pytest.raises(MappingError, match="cycle"):
        topological_sort(entries)


def test_topological_sort_unknown_dep():
    from feasibility.mapping_loader import topological_sort, MappingEntry
    entries = [
        MappingEntry("Details!A1", "black", "a", calc="c", calc_args={"based_on": "missing"}, fallback=0),
    ]
    with pytest.raises(MappingError, match="unknown semantic_name"):
        topological_sort(entries)


def test_validate_against_workbook_missing_cell(tmp_path):
    from feasibility.mapping_loader import validate_against_workbook, MappingFile, MappingEntry
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    wb.active.title = "Details"
    ws = wb["Details"]
    ws["A1"].value = "x"
    ws["A1"].fill = PatternFill("solid", fgColor="FFFFFF00")

    mf = MappingFile(template="x", scheme="s", cells=[
        MappingEntry("Details!A1", "yellow", "a", from_="x.y", fallback=0),
        MappingEntry("Details!B2", "yellow", "b", from_="x.y", fallback=0),  # missing cell
    ])
    with pytest.raises(MappingError, match="Details!B2"):
        validate_against_workbook(mf, wb)


def test_validate_against_workbook_kind_mismatch(tmp_path):
    from feasibility.mapping_loader import validate_against_workbook, MappingFile, MappingEntry
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    wb.active.title = "Details"
    ws = wb["Details"]
    ws["A1"].fill = PatternFill("solid", fgColor="FF000000")  # actually black

    mf = MappingFile(template="x", scheme="s", cells=[
        MappingEntry("Details!A1", "yellow", "a", from_="x.y", fallback=0),
    ])
    with pytest.raises(MappingError, match="kind mismatch"):
        validate_against_workbook(mf, wb)
