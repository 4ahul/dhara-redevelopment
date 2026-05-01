import os

import openpyxl


def verify_report(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Analyzing report: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Check Details sheet
    if "Details" in wb.sheetnames:
        ws = wb["Details"]
        print("\n--- Sheet: Details ---")
        critical_cells = {
            "P5": "Plot Area (sqm)",
            "R19": "Road Width (m)",
            "P44": "RR Residential Rate",
            "J61": "RR Land Rate",
            "Q53": "Residential Carpet Area",
            "P55": "No. of Flats",
            "P61": "Sale Rate",
            "Q56": "Rent Residential",
            "Q57": "Corpus Residential",
            "Q58": "Brokerage Residential",
        }
        for cell, label in critical_cells.items():
            val = ws[cell].value
            print(f"{label} ({cell}): {val}")

    # Check for errors in common calculated cells
    print("\n--- Checking for Errors ---")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        error_cells = []
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value in ["#DIV/0!", "#NAME?", "#REF!", "#VALUE!"]:
                        error_cells.append(f"{cell.coordinate} ({cell.value})")
        except Exception:
            pass

        if error_cells:
            print(
                f"Sheet '{sheet_name}': Found {len(error_cells)} errors. Examples: {error_cells[:10]}"
            )
        else:
            print(f"Sheet '{sheet_name}': No errors found.")


if __name__ == "__main__":
    report_file = (
        "/tmp/reports/Feasibility_33(20)(B)_CLUBBING_Dhiraj_Kunj,_Dwaraka_&_Kalpana_8D001ED7.xlsx"
    )
    verify_report(report_file)
