import openpyxl

from services.template_service import template_service

wb = openpyxl.load_workbook("templates/FINAL TEMPLATE _ 33 (7)(B) .xlsx", data_only=False)
ws = wb["Details"]
for row in ws.iter_rows(min_row=44, max_row=55):
    for cell in row:
        if cell.value and "Corpus" in str(cell.value):
            for x in range(1, 30):
                sibling = ws.cell(row=cell.row, column=x)
                if template_service._is_yellow_cell(sibling):
                    label = template_service._get_cell_label(ws, sibling.row, sibling.column)
