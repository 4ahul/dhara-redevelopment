import os
import sys
import openpyxl

def main(path: str):
    wb = openpyxl.load_workbook(path, data_only=False)
    checks = [
        ("Details", "B19"),
        ("Details", "G34"),
        ("Details", "G39"),
        ("Details", "J45"),
        ("Details", "J54"),
        ("Construction Cost", "D8"),
        ("Construction Cost", "D12"),
        ("Construction Cost", "D15"),
        ("Construction Cost", "H21"),
        ("SUMMARY 1", "I27"),
        ("SUMMARY 1", "I98"),
        ("SUMMARY 1", "I103"),
        ("Profit & Loss Statement", "D28"),
        ("Profit & Loss Statement", "C30"),
        ("Profit & Loss Statement", "D30"),
        ("MCGM PAYMENTS", "B277"),
    ]
    for sheet, cell in checks:
        if sheet not in wb.sheetnames:
            print(f"Missing sheet: {sheet}")
            continue
        ws = wb[sheet]
        val = ws[cell].value
        print(f"{sheet}!{cell} = {val}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.sep, "tmp", "reports", "Feasibility_33_7B_maptest.xlsx")
    if not os.path.exists(path):
        print(f"Not found: {path}")
        sys.exit(1)
    main(path)

