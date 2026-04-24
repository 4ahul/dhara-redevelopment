import openpyxl
from services.template_service import template_service
wb = openpyxl.load_workbook("templates/FINAL TEMPLATE _ 33 (7)(B) .xlsx", data_only=False)
for sheet in wb.sheetnames[:5]:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Existing Carpet Area":
                print(f"Found at {sheet}!{cell.coordinate}")
                # check cells on this row
                for x in range(1, 30):
                    sibling = ws.cell(row=cell.row, column=x)
                    if template_service._is_yellow_cell(sibling):
                        label = template_service._get_cell_label(ws, sibling.row, sibling.column)
                        print(f"Yellow cell at {sibling.coordinate} got label: '{label}'")
