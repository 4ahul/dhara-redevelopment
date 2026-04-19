#!/usr/bin/env python3
import os
import random
import sys
from pathlib import Path

# Add project root and services to path for execution
parent_dir = Path(__file__).parent.parent.parent.absolute()
sys.path.append(str(parent_dir))
sys.path.append(str(Path(__file__).parent.absolute()))

import openpyxl
from services.template_service import template_service, TemplateField
from core.config import settings

def generate_realistic_dummy(label: str, current_value: any) -> any:
    """Generates realistic random dummy data strictly based on labels or existing value types."""
    label_lower = label.lower() if label else ""
    
    # Lead's directed feature test
    if "setback area" in label_lower:
        return 2000
        
    # Text generation
    if isinstance(current_value, str) and not current_value.isdigit() and "area" not in label_lower and "rate" not in label_lower:
        if "name" in label_lower or "developer" in label_lower:
            return f"Dummy Corp {random.randint(10, 99)}"
        if "location" in label_lower or "zone" in label_lower:
            return random.choice(["North", "South West", "Central", "Island City"])
        return f"Test_{random.randint(100, 999)}"

    # Default logic for generic types if value is already a numerical float or int
    if "ratio" in label_lower or "multiplier" in label_lower:
        return round(random.uniform(0.1, 2.5), 2)
    elif "area" in label_lower or "sqm" in label_lower or "sqft" in label_lower or "bua" in label_lower:
        return round(random.uniform(500.0, 15000.0), 2)
    elif "rate" in label_lower or "price" in label_lower or "cost" in label_lower or "premium" in label_lower:
        return round(random.uniform(25000, 150000), 2)
    elif "rent" in label_lower:
        return round(random.uniform(80, 450), 2)
    elif "month" in label_lower or "period" in label_lower:
        return int(random.uniform(12, 60))
    elif "flats" in label_lower or "count" in label_lower or "number" in label_lower:
        return random.randint(10, 200)
    elif "noc" in label_lower or "toggle" in label_lower or "crz" in label_lower:
        return random.choice([0, 1])
    
    # Generic numbers fallback
    return round(random.uniform(10.0, 5000.0), 2)

def main():
    target_filename = "FINAL TEMPLATE _ 33 (7)(B) .xlsx"
    template_path = settings.TEMPLATES_DIR / target_filename
    
    if not template_path.exists():
        print(f"[FAIL] Could not find the template at: {template_path}")
        return
        
    print(f"Loading Template: {template_path.name}...\n")
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    # Limit scope up to P&L if available
    try:
        pl_idx = wb.sheetnames.index("Profit & Loss Statement")
        sheets_to_process = wb.sheetnames[: pl_idx + 1]
    except ValueError:
        sheets_to_process = wb.sheetnames[:5]

    yellow_fields = []
    
    print(f"{'SHEET':<25} | {'CELL':<6} | {'LABEL':<40} | {'ORIGINAL':<12} | {'NEW DUMMY':<12}")
    print("-" * 105)
    
    # Track overrides to feed
    field_overrides = {}

    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if template_service._is_yellow_cell(cell):
                    # Check if it calculates / is formula 
                    cv = str(cell.value).strip() if cell.value else ""
                    if cv.startswith("="):
                        continue # Skip formulas entirely

                    label = template_service._get_cell_label(ws, cell.row, cell.column)
                    original = cell.value
                    
                    # Generate Mock
                    dummy_val = generate_realistic_dummy(label, original)
                    
                    # Record
                    yellow_fields.append((sheet_name, cell.coordinate, label, original, dummy_val))
                    field_overrides[f"{sheet_name}!{cell.coordinate}"] = dummy_val
                    
                    print(f"{sheet_name:<25} | {cell.coordinate:<6} | {str(label)[:38]:<40} | {str(original)[:10]:<12} | {str(dummy_val):<12}")

    print(f"\nExtracted {len(yellow_fields)} input fields.")
    print("Feeding dummy values back into accurate positional cells...")
    
    # Feed back to exactly same workbook structure
    for key, val in field_overrides.items():
        sheet_name, coord = key.split('!')
        wb[sheet_name][coord] = val

    out_name = f"output_test_randomized_{target_filename}"
    out_path = Path(__file__).parent / out_name
    
    wb.save(out_path)
    print(f"[SUCCESS] Randomized test excel generated exactly at: {out_path.absolute()}")

if __name__ == "__main__":
    main()
