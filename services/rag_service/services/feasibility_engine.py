#!/usr/bin/env python3
"""
Intelligent Feasibility Engine
Analyzes property cards and recommends best DCPR clauses with RAG-powered reasoning
"""

import os
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
from dataclasses import dataclass, asdict

env_file = Path(__file__).parent / ".env"
if env_file.exists():
    for line in env_file.read_text().strip().split("\n"):
        if "=" in line:
            key, val = line.split("=", 1)
            os.environ[key] = val

try:
    import easyocr

    EASYOCR_AVAILABLE = True
except ImportError:
    EASYOCR_AVAILABLE = False


@dataclass
class PropertyDetails:
    """Extracted property details"""

    survey_no: str = ""
    plot_area_sq_m: float = 0.0
    plot_area_sq_ft: float = 0.0
    road_width_m: float = 0.0
    zone_type: str = "Residential"
    village: str = ""
    taluka: str = ""
    district: str = ""
    latitude: float = 0.0
    longitude: float = 0.0
    amenity_open_space: float = 0.0
    dpr_remarks: str = ""


@dataclass
class ClauseRecommendation:
    """RAG-based clause recommendation"""

    clause_id: str
    clause_title: str
    relevance_score: float
    summary: str
    applicability: str
    reasoning: str
    conditions: List[str]
    benefits: List[str]


@dataclass
class SchemeAnalysis:
    """Analysis of a DCPR scheme"""

    scheme_id: str
    basic_fsi: float
    incentive_fsi: float
    premium_fsi: float
    max_fsi: float
    total_fsi: float
    is_applicable: bool
    applicability_reason: str
    total_bua_sqft: float
    rehab_area_sqft: float
    saleable_area_sqft: float
    premium_cost: float


@dataclass
class FeasibilityReport:
    """Complete feasibility report"""

    report_id: str
    generated_at: str
    property: PropertyDetails
    best_scheme: str
    best_clauses: List[ClauseRecommendation]
    all_schemes: Dict[str, SchemeAnalysis]
    financial_summary: Dict
    recommendations: List[str]
    reasoning_chain: List[str]
    next_steps: List[str]


class IntelligentOCR:
    """OCR with enhanced extraction using AI"""

    def __init__(self, use_gpu: bool = False):
        self.reader = None
        if EASYOCR_AVAILABLE:
            self.reader = easyocr.Reader(["en"], gpu=use_gpu)
            print("  [OCR] EasyOCR initialized")

    def _extract_with_ai(self, raw_text: str) -> PropertyDetails:
        """Use LLM to extract property details from OCR text"""
        from openai import OpenAI

        client = OpenAI()

        prompt = f"""Extract property details from this OCR text of a property card. Return ONLY valid JSON with these fields:
- survey_no: Survey/Plot number (string)
- plot_area_sq_m: Area in square meters (number)
- road_width_m: Road width in meters (number)
- zone_type: Zone type - Residential/Commercial/Industrial (string)
- village: Village name (string, or empty)
- taluka: Taluka name (string, or empty)
- district: District name (string, or empty)
- latitude: Latitude if present (number, or 0)
- longitude: Longitude if present (number, or 0)

OCR Text:
{raw_text}

Return ONLY the JSON, no other text."""

        try:
            response = client.chat.completions.create(
                model=os.environ.get("OPENAI_MODEL", "llama-3.3-70b-versatile"),
                messages=[
                    {
                        "role": "system",
                        "content": "You are a property data extraction assistant. Extract structured data from OCR text.",
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=500,
            )

            import json as json_module

            result_text = response.choices[0].message.content.strip()
            # Remove markdown code blocks if present
            if result_text.startswith("```"):
                result_text = result_text.split("```")[1]
                if result_text.startswith("json"):
                    result_text = result_text[4:]

            data = json_module.loads(result_text)

            prop = PropertyDetails()
            prop.survey_no = data.get("survey_no", "")
            prop.plot_area_sq_m = float(data.get("plot_area_sq_m", 0))
            prop.road_width_m = float(data.get("road_width_m", 0))
            prop.zone_type = data.get("zone_type", "Residential")
            prop.village = data.get("village", "")
            prop.taluka = data.get("taluka", "")
            prop.district = data.get("district", "")
            prop.latitude = float(data.get("latitude", 0))
            prop.longitude = float(data.get("longitude", 0))

            if prop.plot_area_sq_m > 0:
                prop.plot_area_sq_ft = prop.plot_area_sq_m * 10.764

            return prop
        except Exception as e:
            print(f"  [AI Extraction] Failed: {e}, falling back to regex")
            return None

    def extract_from_file(self, file_path: str) -> PropertyDetails:
        """Extract property details from PDF or image"""
        from property_card_workflow import PropertyCardOCR

        if file_path.lower().endswith(".pdf"):
            card = PropertyCardOCR()
            cards = card.extract_from_pdf(file_path)
            if cards:
                return self._card_to_property(cards[0])
        else:
            result = self.reader.readtext(file_path)

            # Build raw text from OCR
            raw_text = "\n".join([r[1] for r in result])
            print("  [OCR] Raw text extracted, using AI for parsing...")

            # Try AI extraction first
            prop = self._extract_with_ai(raw_text)
            if prop and prop.survey_no:
                return prop

            # Fallback to regex extraction
            print("  [OCR] Falling back to regex extraction...")
            return self._extract_from_ocr_result(result)

    def _card_to_property(self, card) -> PropertyDetails:
        """Convert PropertyCard to PropertyDetails"""
        prop = PropertyDetails()
        prop.survey_no = card.survey_no
        prop.plot_area_sq_m = card.plot_area_sq_m
        prop.plot_area_sq_ft = card.plot_area_sq_ft
        prop.road_width_m = card.road_width_m
        prop.zone_type = card.zone_type
        prop.village = card.village
        prop.taluka = card.taluka
        prop.district = card.district
        return prop

    def _extract_from_ocr_result(self, result) -> PropertyDetails:
        """Extract from OCR result - improved extraction"""
        prop = PropertyDetails()

        # Build text blocks with positions
        blocks = []
        for detection in result:
            bbox = detection[0]
            text = detection[1]
            # Get x position (left edge)
            x_pos = min([p[0] for p in bbox])
            y_pos = min([p[1] for p in bbox])
            blocks.append(
                {
                    "text": text,
                    "text_upper": text.upper(),
                    "x": x_pos,
                    "y": y_pos,
                    "bbox": bbox,
                }
            )

        # Pair labels with values based on position
        for i, block in enumerate(blocks):
            text = block["text_upper"]
            text_lower = block["text"].lower()

            # Survey No
            if "SURVEY" in text or "SY NO" in text or "CTS" in text:
                match = re.search(r"(\d+[\/\-\d]*)", block["text"])
                if match:
                    prop.survey_no = match.group(1)
                # Look for value to the right
                for j in range(i + 1, len(blocks)):
                    if (
                        abs(blocks[j]["x"] - block["x"]) < 200
                        and abs(blocks[j]["y"] - block["y"]) < 20
                    ):
                        val_match = re.search(r"(\d+[\/\-\d]*)", blocks[j]["text"])
                        if val_match and not prop.survey_no:
                            prop.survey_no = val_match.group(1)
                        break

            # Village
            if "VILLAGE" in text or "VILL" in text:
                for j in range(i + 1, len(blocks)):
                    if (
                        abs(blocks[j]["x"] - block["x"]) < 200
                        and abs(blocks[j]["y"] - block["y"]) < 20
                    ):
                        val = blocks[j]["text"].strip()
                        if val and not any(c.isdigit() for c in val):
                            prop.village = val
                        break

            # Taluka
            if "TALUKA" in text:
                for j in range(i + 1, len(blocks)):
                    if (
                        abs(blocks[j]["x"] - block["x"]) < 200
                        and abs(blocks[j]["y"] - block["y"]) < 20
                    ):
                        val = blocks[j]["text"].strip()
                        if val and not any(c.isdigit() for c in val):
                            prop.taluka = val
                        break

            # District
            if "DISTRICT" in text:
                for j in range(i + 1, len(blocks)):
                    if (
                        abs(blocks[j]["x"] - block["x"]) < 200
                        and abs(blocks[j]["y"] - block["y"]) < 20
                    ):
                        val = blocks[j]["text"].strip()
                        if val:
                            prop.district = val
                        break

            # Plot Area
            if ("AREA" in text or "PLOT" in text) and "sq" in text_lower:
                match = re.search(r"(\d+[\.\d]*)", block["text"])
                if match:
                    prop.plot_area_sq_m = float(match.group(1))

            # Road Width
            if "ROAD" in text and "WIDTH" in text:
                match = re.search(r"(\d+[\.\d]*)", block["text"])
                if match:
                    prop.road_width_m = float(match.group(1))
                # Look for value nearby
                for j in range(i + 1, len(blocks)):
                    if abs(blocks[j]["y"] - block["y"]) < 20:
                        val_match = re.search(r"(\d+[\.\d]*)", blocks[j]["text"])
                        if val_match and prop.road_width_m == 0:
                            prop.road_width_m = float(val_match.group(1))
                        break

            # Zone
            if "ZONE" in text:
                if "RESIDENTIAL" in text:
                    prop.zone_type = "Residential"
                elif "COMMERCIAL" in text:
                    prop.zone_type = "Commercial"
                elif "INDUSTRIAL" in text:
                    prop.zone_type = "Industrial"
                else:
                    for j in range(i + 1, len(blocks)):
                        if (
                            abs(blocks[j]["x"] - block["x"]) < 200
                            and abs(blocks[j]["y"] - block["y"]) < 20
                        ):
                            val = blocks[j]["text"].strip()
                            if "RESIDENTIAL" in val.upper():
                                prop.zone_type = "Residential"
                            elif "COMMERCIAL" in val.upper():
                                prop.zone_type = "Commercial"
                            elif "INDUSTRIAL" in val.upper():
                                prop.zone_type = "Industrial"
                            break

            # Latitude
            if "LATITUDE" in text:
                match = re.search(r"([\d.]+)", block["text"])
                if match:
                    prop.latitude = float(match.group(1))

            # Longitude
            if "LONGITUDE" in text:
                match = re.search(r"([\d.]+)", block["text"])
                if match:
                    prop.longitude = float(match.group(1))

        if prop.plot_area_sq_m > 0:
            prop.plot_area_sq_ft = prop.plot_area_sq_m * 10.764

        return prop


class DCPRClauseFinder:
    """RAG-powered clause finder"""

    def __init__(self):
        self._agent = None
        self._vectorstore = None

    @property
    def agent(self):
        if self._agent is None:
            from rag import RAGAgent

            self._agent = RAGAgent(use_milvus=True)
        return self._agent

    @property
    def vectorstore(self):
        if self._vectorstore is None:
            self._vectorstore = self.agent.vectorstore
        return self._vectorstore

    def find_applicable_clauses(
        self, prop: PropertyDetails
    ) -> List[ClauseRecommendation]:
        """Find best DCPR clauses based on property details"""
        queries = self._build_queries(prop)
        all_results = []

        for query in queries:
            results = self.vectorstore.search(query, k=5)
            for score, text in results:
                clause = self._extract_clause_info(text, score, prop)
                if clause and not self._is_duplicate(clause, all_results):
                    all_results.append(clause)

        all_results.sort(key=lambda x: x.relevance_score, reverse=True)
        return all_results[:10]

    def _build_queries(self, prop: PropertyDetails) -> List[str]:
        """Build targeted queries based on property"""
        queries = [
            f"FSI {prop.zone_type} zone {prop.road_width_m}m road",
            "DCPR 33(7B) 33(20B) residential redevelopment clauses",
            f"open space marginal setbacks {prop.zone_type}",
            f"building permission {prop.zone_type} zone DCPR requirements",
        ]

        if prop.plot_area_sq_m < 4000:
            queries.append(f"FSI table plots under 4000 sqm {prop.road_width_m}m")
        if prop.road_width_m >= 12:
            queries.append("FSI 12 meter road width residential maximum")
        if prop.zone_type == "Residential":
            queries.append("residential zone DCPR clause building requirements")

        return queries

    def _extract_clause_info(
        self, text: str, score: float, prop: PropertyDetails
    ) -> Optional[ClauseRecommendation]:
        """Extract and analyze clause information"""
        clause_match = re.search(r"Clause\s*(\d+(?:\([\w]+\))?)", text, re.IGNORECASE)
        table_match = re.search(r"Table\s*No\.?\s*(\d+[a-zA-Z]?)", text, re.IGNORECASE)

        clause_id = clause_match.group(1) if clause_match else "General"
        table_no = table_match.group(1) if table_match else None

        applicability = self._determine_applicability(text, prop)
        if not applicability:
            return None

        reasoning = self._generate_reasoning(text, prop, clause_id)
        conditions = self._extract_conditions(text)
        benefits = self._extract_benefits(text)

        return ClauseRecommendation(
            clause_id=clause_id,
            clause_title=self._get_clause_title(clause_id, text),
            relevance_score=float(score),
            summary=text[:400],
            applicability=applicability,
            reasoning=reasoning,
            conditions=conditions,
            benefits=benefits,
        )

    def _determine_applicability(self, text: str, prop: PropertyDetails) -> str:
        """Determine if clause applies to property"""
        text_lower = text.lower()
        zone_lower = prop.zone_type.lower()

        if "residential" in text_lower and zone_lower == "residential":
            return f"Applicable to {prop.zone_type} zone"
        if "commercial" in text_lower and zone_lower == "commercial":
            return f"Applicable to {prop.zone_type} zone"
        if "all zones" in text_lower:
            return "Universal application"

        if any(term in text_lower for term in ["redevelopment", "society", "housing"]):
            if prop.zone_type == "Residential":
                return f"Potentially applicable for {prop.zone_type} redevelopment"

        return "May be applicable"

    def _generate_reasoning(
        self, text: str, prop: PropertyDetails, clause_id: str
    ) -> str:
        """Generate reasoning for clause selection"""
        reasons = []

        if "33(7B)" in text or "33(7)" in clause_id:
            reasons.append(
                "33(7B) provides basic FSI of 0.5 for rehabilitation-focused redevelopment"
            )
        if "33(20B)" in text or "33(20)" in clause_id:
            reasons.append(
                "33(20B) offers higher basic FSI of 2.5 for residential areas"
            )
        if "incentive" in text.lower():
            reasons.append("Contains incentive FSI provisions for eligible projects")
        if "premium" in text.lower():
            reasons.append("Includes premium FSI options for additional development")

        if prop.road_width_m >= 12:
            reasons.append(
                f"Road width of {prop.road_width_m}m qualifies for enhanced FSI under Table 12"
            )

        if prop.plot_area_sq_m < 4000:
            reasons.append(
                f"Plot area of {prop.plot_area_sq_m} sq.m falls under smaller plot category"
            )

        return ". ".join(reasons) if reasons else "Standard building requirements apply"

    def _extract_conditions(self, text: str) -> List[str]:
        """Extract conditions from text"""
        conditions = []
        text_lower = text.lower()

        if "70%" in text or "seventy" in text_lower:
            conditions.append("70% affordable housing requirement")
        if "consent" in text_lower:
            conditions.append("Requires 70% society member consent")
        if "rehabilitation" in text_lower:
            conditions.append("Must provide rehabilitation component")
        if "premium" in text_lower:
            conditions.append("Premium payment may be required")

        return conditions[:5]

    def _extract_benefits(self, text: str) -> List[str]:
        """Extract benefits from text"""
        benefits = []
        text_lower = text.lower()

        if "fsi" in text_lower:
            if "2.5" in text:
                benefits.append("Higher FSI potential (up to 2.5)")
            if "3.5" in text:
                benefits.append("Enhanced FSI available (up to 3.5)")
            if "4.0" in text or "4.00" in text:
                benefits.append("Maximum FSI of 4.0 available")
        if "rehabilitation" in text_lower:
            benefits.append("Rehabilitation component benefits")
        if "incentive" in text_lower:
            benefits.append("Incentive FSI for qualifying projects")

        return benefits[:5]

    def _get_clause_title(self, clause_id: str, text: str) -> str:
        """Get clause title"""
        titles = {
            "33(7B)": "Redevelopment of Existing Residential Premises",
            "33(20B)": "Redevelopment of Residential/Commercial Buildings",
            "33(11)": "Permanent Transit Camp Tenements",
            "30(A)": "Transit Oriented Development",
            "12": "FSI Table - Maximum Permissible",
        }

        for key, title in titles.items():
            if key in clause_id or key in text[:200]:
                return title

        return f"DCPR Clause {clause_id}"

    def _is_duplicate(
        self, clause: ClauseRecommendation, existing: List[ClauseRecommendation]
    ) -> bool:
        """Check if clause is duplicate"""
        for e in existing:
            if e.clause_id == clause.clause_id:
                return True
        return False


class SchemeCalculator:
    """Calculate DCPR schemes dynamically"""

    def __init__(self):
        self.schemes = {
            "33(7B)": {
                "basic_fsi": 0.5,
                "incentive_options": {"70%": 0.15, "35%": 0.10, "20%": 0.05},
                "max_fsi": 4.0,
                "description": "Rehabilitation-focused redevelopment",
            },
            "33(20B)": {
                "basic_fsi": 2.5,
                "incentive_options": {},
                "max_fsi": 4.0,
                "description": "Standard residential redevelopment",
            },
            "33(11)": {
                "basic_fsi": 1.0,
                "incentive_options": {},
                "max_fsi": 4.0,
                "description": "Transit camp tenements",
            },
            "30(A)": {
                "basic_fsi": 2.5,
                "incentive_options": {},
                "max_fsi": 4.0,
                "description": "Transit Oriented Development",
            },
        }

    def get_fsi_table_fsi(self, area_sqm: float, road_width: float) -> float:
        """Get FSI from DCPR Table 12"""
        if area_sqm <= 4000:
            if road_width >= 27:
                return 3.5
            elif road_width >= 18:
                return 3.0
            elif road_width >= 12:
                return 2.5
            elif road_width >= 9:
                return 2.25
        elif area_sqm <= 10000:
            if road_width >= 27:
                return 5.0
            elif road_width >= 18:
                return 4.0
            elif road_width >= 12:
                return 3.5
            elif road_width >= 9:
                return 2.75
        elif area_sqm <= 20000:
            if road_width >= 27:
                return 6.5
            elif road_width >= 18:
                return 5.0
            elif road_width >= 12:
                return 4.0
            elif road_width >= 9:
                return 3.5
        return 2.5

    def get_marginal_distances(
        self, area_sqm: float, road_width: float, height: float = 15.0
    ) -> Dict[str, float]:
        """Get marginal open spaces (setbacks) per Pune DCPR 2017"""
        # Simplified Pune Residential Setbacks
        front_margin = 3.0
        if road_width >= 18:
            front_margin = 6.0
        elif road_width >= 12:
            front_margin = 4.5

        side_margin = 3.0
        if height > 15:
            side_margin = max(3.0, height / 4.0)

        return {"front": front_margin, "side": side_margin, "rear": side_margin}

    def analyze_scheme(
        self, scheme_id: str, prop: PropertyDetails, affordable_pct: float = 70
    ) -> SchemeAnalysis:
        """Analyze a specific scheme"""
        scheme = self.schemes.get(scheme_id, self.schemes["33(20B)"])

        basic_fsi = scheme["basic_fsi"]
        incentive_fsi = 0.0

        if scheme_id == "33(7B)" and affordable_pct in scheme["incentive_options"]:
            incentive_fsi = scheme["incentive_options"][affordable_pct]

        total_fsi = min(basic_fsi + incentive_fsi, scheme["max_fsi"])

        is_applicable = self._check_applicability(scheme_id, prop)
        applicability_reason = self._get_applicability_reason(scheme_id, prop)

        total_bua = prop.plot_area_sq_ft * total_fsi
        rehab_area = total_bua * 0.7
        saleable_area = total_bua * 0.3

        premium_cost = self._calculate_premium(prop, total_fsi)

        return SchemeAnalysis(
            scheme_id=scheme_id,
            basic_fsi=basic_fsi,
            incentive_fsi=incentive_fsi,
            premium_fsi=0.5,
            max_fsi=scheme["max_fsi"],
            total_fsi=total_fsi,
            is_applicable=is_applicable,
            applicability_reason=applicability_reason,
            total_bua_sqft=total_bua,
            rehab_area_sqft=rehab_area,
            saleable_area_sqft=saleable_area,
            premium_cost=premium_cost,
        )

    def _check_applicability(self, scheme_id: str, prop: PropertyDetails) -> bool:
        """Check if scheme applies to property"""
        if scheme_id == "33(7B)":
            return prop.zone_type == "Residential"
        if scheme_id == "33(20B)":
            return prop.zone_type in ["Residential", "Commercial"]
        if scheme_id == "33(11)":
            return prop.zone_type == "Residential"
        if scheme_id == "30(A)":
            return prop.road_width_m >= 12
        return True

    def _get_applicability_reason(self, scheme_id: str, prop: PropertyDetails) -> str:
        """Get reason for applicability"""
        reasons = {
            "33(7B)": "Residential zone qualifies for rehabilitation-focused redevelopment",
            "33(20B)": f"Standard redevelopment clause for {prop.zone_type} buildings",
            "33(11)": "For permanent transit camp tenements",
            "30(A)": f"Road width {prop.road_width_m}m meets TOD requirements",
        }
        return reasons.get(scheme_id, "Standard applicability")

    def _calculate_premium(self, prop: PropertyDetails, total_fsi: float) -> float:
        """Estimate premium cost"""
        base_premium = prop.plot_area_sq_ft * 500
        fsi_premium = (
            (total_fsi - 2.5) * prop.plot_area_sq_ft * 1000 if total_fsi > 2.5 else 0
        )
        return base_premium + fsi_premium


class FeasibilityEngine:
    """
    Main intelligent feasibility engine
    Combines OCR, RAG clause finding, and scheme analysis
    """

    def __init__(self):
        self.ocr = IntelligentOCR()
        self.clause_finder = DCPRClauseFinder()
        self.calculator = SchemeCalculator()

    def analyze_from_file(
        self, file_path: str, affordable_pct: float = 70
    ) -> FeasibilityReport:
        """
        Complete feasibility analysis from property card file

        Args:
            file_path: Path to property card PDF/image
            affordable_pct: Affordable housing percentage (for 33(7B))

        Returns:
            FeasibilityReport with full analysis and recommendations
        """
        print(f"\n{'=' * 60}")
        print("INTELLIGENT FEASIBILITY ANALYSIS")
        print(f"{'=' * 60}\n")

        print("[1/5] Extracting property details from file...")
        prop = self.ocr.extract_from_file(file_path)
        if not prop.survey_no:
            prop.survey_no = "EXTRACTED_FROM_FILE"
        print(f"  Survey No: {prop.survey_no}")
        print(
            f"  Area: {prop.plot_area_sq_m:.0f} sq.m ({prop.plot_area_sq_ft:.0f} sq.ft)"
        )
        print(f"  Road Width: {prop.road_width_m}m")
        print(f"  Zone: {prop.zone_type}")

        print("\n[2/5] Searching DCPR for applicable clauses (RAG)...")
        clauses = self.clause_finder.find_applicable_clauses(prop)
        print(f"  Found {len(clauses)} relevant clauses")

        print("\n[3/5] Analyzing DCPR schemes...")
        schemes = {}
        for scheme_id in ["33(7B)", "33(20B)", "33(11)", "30(A)"]:
            analysis = self.calculator.analyze_scheme(scheme_id, prop, affordable_pct)
            schemes[scheme_id] = analysis
            status = "APPLICABLE" if analysis.is_applicable else "NOT APPLICABLE"
            print(
                f"  {scheme_id}: Basic {analysis.basic_fsi} + Incentive {analysis.incentive_fsi} = Total {analysis.total_fsi} [{status}]"
            )

        print("\n[4/5] Generating recommendations...")
        best_scheme = self._find_best_scheme(schemes, prop)
        recommendations = self._generate_recommendations(
            prop, schemes, clauses, best_scheme
        )
        reasoning = self._generate_reasoning_chain(prop, schemes, clauses, best_scheme)
        next_steps = self._generate_next_steps(prop, best_scheme, clauses)

        print("\n[5/5] Building financial summary...")
        financial = self._build_financial_summary(prop, schemes[best_scheme])

        report = FeasibilityReport(
            report_id=f"FEAS_{prop.survey_no.replace('/', '_')}_{datetime.now().strftime('%Y%m%d%H%M')}",
            generated_at=datetime.now().isoformat(),
            property=prop,
            best_scheme=best_scheme,
            best_clauses=clauses[:5],
            all_schemes={k: self._scheme_to_dict(v) for k, v in schemes.items()},
            financial_summary=financial,
            recommendations=recommendations,
            reasoning_chain=reasoning,
            next_steps=next_steps,
        )

        print(f"\n{'=' * 60}")
        print(f"BEST SCHEME: {best_scheme}")
        print(f"TOTAL FSI: {schemes[best_scheme].total_fsi}")
        print(
            f"RECOMMENDATION: {recommendations[0] if recommendations else 'See full report'}"
        )
        print(f"{'=' * 60}\n")

        return report

    def analyze_from_params(
        self,
        survey_no: str,
        area_sqm: float,
        road_width: float,
        zone: str,
        affordable_pct: float = 70,
    ) -> FeasibilityReport:
        """Analyze from direct parameters"""
        prop = PropertyDetails(
            survey_no=survey_no,
            plot_area_sq_m=area_sqm,
            plot_area_sq_ft=area_sqm * 10.764,
            road_width_m=road_width,
            zone_type=zone,
        )
        file_path = "PARAMETERS_ONLY"
        return self._run_analysis(prop, affordable_pct)

    def _run_analysis(
        self, prop: PropertyDetails, affordable_pct: float
    ) -> FeasibilityReport:
        """Internal analysis runner"""
        clauses = self.clause_finder.find_applicable_clauses(prop)
        schemes = {}

        for scheme_id in ["33(7B)", "33(20B)", "33(11)", "30(A)"]:
            schemes[scheme_id] = self.calculator.analyze_scheme(
                scheme_id, prop, affordable_pct
            )

        best_scheme = self._find_best_scheme(schemes, prop)
        recommendations = self._generate_recommendations(
            prop, schemes, clauses, best_scheme
        )
        reasoning = self._generate_reasoning_chain(prop, schemes, clauses, best_scheme)
        next_steps = self._generate_next_steps(prop, best_scheme, clauses)
        financial = self._build_financial_summary(prop, schemes[best_scheme])

        return FeasibilityReport(
            report_id=f"FEAS_{prop.survey_no.replace('/', '_')}_{datetime.now().strftime('%Y%m%d%H%M')}",
            generated_at=datetime.now().isoformat(),
            property=prop,
            best_scheme=best_scheme,
            best_clauses=clauses[:5],
            all_schemes={k: self._scheme_to_dict(v) for k, v in schemes.items()},
            financial_summary=financial,
            recommendations=recommendations,
            reasoning_chain=reasoning,
            next_steps=next_steps,
        )

    def _find_best_scheme(
        self, schemes: Dict[str, SchemeAnalysis], prop: PropertyDetails
    ) -> str:
        """Find best scheme based on FSI and conditions"""
        applicable = {k: v for k, v in schemes.items() if v.is_applicable}

        if not applicable:
            return "33(20B)"

        best = max(
            applicable.items(), key=lambda x: (x[1].total_fsi, -x[1].premium_cost)
        )
        return best[0]

    def _generate_recommendations(
        self,
        prop: PropertyDetails,
        schemes: Dict[str, SchemeAnalysis],
        clauses: List[ClauseRecommendation],
        best_scheme: str,
    ) -> List[str]:
        """Generate recommendations with reasoning"""
        recs = []
        best = schemes[best_scheme]

        recs.append(
            f"Recommended scheme: {best_scheme} with total FSI of {best.total_fsi}"
        )

        if best_scheme == "33(7B)":
            recs.append(
                "33(7B) provides 0.5 basic FSI + 0.15 incentive FSI for rehabilitation"
            )
            recs.append("Must allocate 70% of BUA for rehabilitation component")
        elif best_scheme == "33(20B)":
            recs.append("33(20B) offers 2.5 basic FSI for standard redevelopment")
            recs.append(
                "No mandatory rehabilitation component if building is not cessed"
            )

        if prop.road_width_m >= 12:
            recs.append(
                f"Road width {prop.road_width_m}m qualifies for enhanced FSI per Table 12"
            )

        if clauses:
            top_clauses = [c.clause_id for c in clauses[:3] if c.clause_id != "General"]
            if top_clauses:
                recs.append(f"Key DCPR clauses to reference: {', '.join(top_clauses)}")

        return recs

    def _generate_reasoning_chain(
        self,
        prop: PropertyDetails,
        schemes: Dict[str, SchemeAnalysis],
        clauses: List[ClauseRecommendation],
        best_scheme: str,
    ) -> List[str]:
        """Generate step-by-step reasoning"""
        chain = []

        chain.append(
            f"1. Property analysis: {prop.survey_no}, {prop.plot_area_sq_m:.0f} sq.m, {prop.zone_type} zone"
        )

        chain.append(
            f"2. Road width {prop.road_width_m}m meets minimum requirements for DCPR Table 12 FSI"
        )

        scheme_scores = []
        for scheme_id, scheme in schemes.items():
            if scheme.is_applicable:
                score = scheme.total_fsi * 100 - scheme.premium_cost / 100000
                scheme_scores.append((scheme_id, scheme.total_fsi, score))
                chain.append(
                    f"3.{len(scheme_scores)}. {scheme_id}: Basic {scheme.basic_fsi} + Incentive {scheme.incentive_fsi} = Total {scheme.total_fsi}"
                )

        chain.append(
            f"4. Selected {best_scheme} as it provides highest total FSI of {schemes[best_scheme].total_fsi}"
        )

        if clauses:
            chain.append(
                f"5. RAG analysis identified {len(clauses)} relevant clauses from DCPR 2034"
            )
            for c in clauses[:3]:
                chain.append(f"   - Clause {c.clause_id}: {c.reasoning[:80]}...")

        return chain

    def _generate_next_steps(
        self,
        prop: PropertyDetails,
        best_scheme: str,
        clauses: List[ClauseRecommendation],
    ) -> List[str]:
        """Generate next steps"""
        steps = [
            f"1. Engage architect to prepare plans under {best_scheme}",
            "2. Obtain society resolution with 70% consent",
            "3. Submit application to MCGM with property documents",
        ]

        if clauses:
            steps.append("4. Reference applicable DCPR clauses in application")
            steps.append("5. Prepare documentation per identified clauses")

        steps.extend(
            [
                "6. Obtain provisional approval (IOD)",
                "7. Commence construction after all NOCs",
                "8. Apply for occupation certificate",
            ]
        )

        return steps

    def _build_financial_summary(
        self, prop: PropertyDetails, scheme: SchemeAnalysis
    ) -> Dict:
        """Build financial summary"""
        rate_per_sqft = 25000
        saleable_value = scheme.saleable_area_sqft * rate_per_sqft

        construction_cost = scheme.total_bua_sqft * 2500
        premium = scheme.premium_cost
        approval_cost = prop.plot_area_sq_ft * 100

        return {
            "total_bua_sqft": scheme.total_bua_sqft,
            "rehab_area_sqft": scheme.rehab_area_sqft,
            "saleable_area_sqft": scheme.saleable_area_sqft,
            "estimated_value_cr": round(saleable_value / 10000000, 2),
            "construction_cost_cr": round(construction_cost / 10000000, 2),
            "premium_cost_cr": round(premium / 10000000, 2),
            "approval_cost_cr": round(approval_cost / 10000000, 2),
            "rate_per_sqft": rate_per_sqft,
        }

    def _scheme_to_dict(self, scheme: SchemeAnalysis) -> Dict:
        """Convert scheme to dict"""
        return {
            "basic_fsi": scheme.basic_fsi,
            "incentive_fsi": scheme.incentive_fsi,
            "premium_fsi": scheme.premium_fsi,
            "max_fsi": scheme.max_fsi,
            "total_fsi": scheme.total_fsi,
            "is_applicable": scheme.is_applicable,
            "applicability_reason": scheme.applicability_reason,
            "total_bua_sqft": scheme.total_bua_sqft,
            "rehab_area_sqft": scheme.rehab_area_sqft,
            "saleable_area_sqft": scheme.saleable_area_sqft,
            "premium_cost": scheme.premium_cost,
        }

    def export_to_json(self, report: FeasibilityReport, output_path: str):
        """Export report to JSON"""
        data = asdict(report)
        with open(output_path, "w") as f:
            json.dump(data, f, indent=2)
        print(f"  Exported to: {output_path}")

    def export_to_excel(self, report: FeasibilityReport, output_path: str):
        """Export report to Excel with multiple sheets"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="2C3E50", end_color="2C3E50", fill_type="solid"
        )
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(
            start_color="3498DB", end_color="3498DB", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def style_header(cell, header=False, subheader=False):
            if header:
                cell.font = header_font
                cell.fill = header_fill
            elif subheader:
                cell.font = subheader_font
                cell.fill = subheader_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

        # Sheet 1: Property Details
        ws1 = wb.active
        ws1.title = "Property Details"

        data = [
            ["FEASIBILITY REPORT - DCPR ANALYSIS", ""],
            ["Report ID", report.report_id],
            ["Generated", report.generated_at],
            ["", ""],
            ["PROPERTY DETAILS", ""],
            ["Survey No", report.property.survey_no],
            ["Plot Area (sq.m)", f"{report.property.plot_area_sq_m:.2f}"],
            ["Plot Area (sq.ft)", f"{report.property.plot_area_sq_ft:.2f}"],
            ["Road Width (m)", f"{report.property.road_width_m:.2f}"],
            ["Zone Type", report.property.zone_type],
            ["Village", report.property.village or "N/A"],
            ["Taluka", report.property.taluka or "N/A"],
            ["District", report.property.district or "N/A"],
        ]

        for i, row in enumerate(data, 1):
            ws1.cell(row=i, column=1, value=row[0])
            if len(row) > 1:
                ws1.cell(row=i, column=2, value=row[1])
            if i == 1:
                ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
                ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            elif i == 5:
                ws1.cell(row=i, column=1).font = Font(
                    bold=True, size=12, color="FFFFFF"
                )
                ws1.cell(row=i, column=1).fill = PatternFill(
                    start_color="27AE60", end_color="27AE60", fill_type="solid"
                )
                ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

        for col in range(1, 3):
            ws1.column_dimensions[get_column_letter(col)].width = 30

        # Sheet 2: Scheme Comparison
        ws2 = wb.create_sheet("Scheme Comparison")

        headers = [
            "Scheme",
            "Basic FSI",
            "Incentive FSI",
            "Premium FSI",
            "Max FSI",
            "Total FSI",
            "Applicable",
            "Total BUA (sq.ft)",
            "Premium Cost (Cr)",
        ]
        ws2.append(headers)
        for col, _ in enumerate(headers, 1):
            style_header(ws2.cell(1, col), header=True)

        for i, (scheme_id, scheme) in enumerate(report.all_schemes.items(), 2):
            row_data = [
                scheme_id,
                scheme["basic_fsi"],
                scheme["incentive_fsi"],
                scheme["premium_fsi"],
                scheme["max_fsi"],
                scheme["total_fsi"],
                "YES" if scheme["is_applicable"] else "NO",
                f"{scheme['total_bua_sqft']:,.0f}",
                f"{scheme['premium_cost'] / 10000000:.2f}",
            ]
            ws2.append(row_data)
            for col, _ in enumerate(row_data, 1):
                cell = ws2.cell(i, col)
                cell.border = thin_border
                if col == 7:
                    cell.fill = PatternFill(
                        start_color="27AE60" if scheme["is_applicable"] else "E74C3C",
                        end_color="27AE60" if scheme["is_applicable"] else "E74C3C",
                        fill_type="solid",
                    )
                    cell.font = Font(color="FFFFFF", bold=True)
                if i == 2:
                    cell.fill = PatternFill(
                        start_color="F1C40F", end_color="F1C40F", fill_type="solid"
                    )

        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 18

        # Sheet 3: Best Scheme Analysis
        ws3 = wb.create_sheet("Best Scheme")
        best_scheme = report.all_schemes.get(report.best_scheme, {})

        best_data = [
            ["BEST SCHEME RECOMMENDATION", ""],
            ["Recommended Scheme", report.best_scheme],
            ["", ""],
            ["FSI BREAKDOWN", ""],
            ["Basic FSI", best_scheme.get("basic_fsi", 0)],
            ["Incentive FSI", best_scheme.get("incentive_fsi", 0)],
            ["Premium FSI", best_scheme.get("premium_fsi", 0)],
            ["Total FSI", best_scheme.get("total_fsi", 0)],
            ["Maximum FSI", best_scheme.get("max_fsi", 0)],
            ["", ""],
            ["AREA BREAKDOWN", ""],
            ["Total BUA (sq.ft)", f"{best_scheme.get('total_bua_sqft', 0):,.0f}"],
            [
                "Rehabilitation Area (sq.ft)",
                f"{best_scheme.get('rehab_area_sqft', 0):,.0f}",
            ],
            [
                "Saleable Area (sq.ft)",
                f"{best_scheme.get('saleable_area_sqft', 0):,.0f}",
            ],
            ["", ""],
            ["APPLICABILITY", ""],
            ["Is Applicable", "YES" if best_scheme.get("is_applicable") else "NO"],
            ["Reason", best_scheme.get("applicability_reason", "N/A")],
        ]

        for i, row in enumerate(best_data, 1):
            ws3.cell(row=i, column=1, value=row[0])
            ws3.cell(row=i, column=2, value=row[1] if len(row) > 1 else "")
            ws3.cell(i, 1).font = Font(bold=True)
            if "BREAKDOWN" in str(row[0]) or row[0] in [
                "FSI BREAKDOWN",
                "AREA BREAKDOWN",
                "APPLICABILITY",
                "BEST SCHEME RECOMMENDATION",
            ]:
                ws3.cell(i, 1).font = Font(bold=True, color="FFFFFF", size=11)
                ws3.cell(i, 1).fill = PatternFill(
                    start_color="2C3E50", end_color="2C3E50", fill_type="solid"
                )
                ws3.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

        ws3.column_dimensions["A"].width = 35
        ws3.column_dimensions["B"].width = 25

        # Sheet 4: DCPR Clauses
        ws4 = wb.create_sheet("DCPR Clauses")

        clause_headers = [
            "Clause ID",
            "Title",
            "Relevance Score",
            "Applicability",
            "Reasoning",
            "Conditions",
            "Benefits",
        ]
        ws4.append(clause_headers)
        for col, _ in enumerate(clause_headers, 1):
            style_header(ws4.cell(1, col), header=True)

        for i, clause in enumerate(report.best_clauses, 2):
            row_data = [
                clause.clause_id,
                clause.clause_title,
                f"{clause.relevance_score:.3f}",
                clause.applicability,
                clause.reasoning[:100] + "..."
                if len(clause.reasoning) > 100
                else clause.reasoning,
                "; ".join(clause.conditions[:3]) if clause.conditions else "N/A",
                "; ".join(clause.benefits[:3]) if clause.benefits else "N/A",
            ]
            ws4.append(row_data)
            for col, _ in enumerate(row_data, 1):
                ws4.cell(i, col).border = thin_border
                ws4.cell(i, col).alignment = Alignment(wrap_text=True, vertical="top")

        ws4.column_dimensions["A"].width = 12
        ws4.column_dimensions["B"].width = 30
        ws4.column_dimensions["C"].width = 15
        ws4.column_dimensions["D"].width = 25
        ws4.column_dimensions["E"].width = 50
        ws4.column_dimensions["F"].width = 30
        ws4.column_dimensions["G"].width = 30

        # Sheet 5: Financial Summary
        ws5 = wb.create_sheet("Financial Summary")

        fin_headers = ["Item", "Value", "Unit"]
        ws5.append(fin_headers)
        for col, _ in enumerate(fin_headers, 1):
            style_header(ws5.cell(1, col), header=True)

        fin_data = [
            [
                "Total BUA",
                f"{report.financial_summary.get('total_bua_sqft', 0):,.0f}",
                "sq.ft",
            ],
            [
                "Rehabilitation Area",
                f"{report.financial_summary.get('rehab_area_sqft', 0):,.0f}",
                "sq.ft",
            ],
            [
                "Saleable Area",
                f"{report.financial_summary.get('saleable_area_sqft', 0):,.0f}",
                "sq.ft",
            ],
            [
                "Rate per sq.ft",
                f"{report.financial_summary.get('rate_per_sqft', 0):,.0f}",
                "INR",
            ],
            [
                "Estimated Value",
                f"{report.financial_summary.get('estimated_value_cr', 0):.2f}",
                "Crores",
            ],
            [
                "Construction Cost",
                f"{report.financial_summary.get('construction_cost_cr', 0):.2f}",
                "Crores",
            ],
            [
                "Premium Cost",
                f"{report.financial_summary.get('premium_cost_cr', 0):.2f}",
                "Crores",
            ],
            [
                "Approval Cost",
                f"{report.financial_summary.get('approval_cost_cr', 0):.2f}",
                "Crores",
            ],
            ["", "", ""],
            ["PROFITABILITY", "", ""],
            [
                "Gross Value",
                f"{report.financial_summary.get('estimated_value_cr', 0):.2f}",
                "Crores",
            ],
            [
                "Total Costs",
                f"{(report.financial_summary.get('construction_cost_cr', 0) + report.financial_summary.get('premium_cost_cr', 0) + report.financial_summary.get('approval_cost_cr', 0)):.2f}",
                "Crores",
            ],
            [
                "Net Value",
                f"{(report.financial_summary.get('estimated_value_cr', 0) - report.financial_summary.get('construction_cost_cr', 0) - report.financial_summary.get('premium_cost_cr', 0) - report.financial_summary.get('approval_cost_cr', 0)):.2f}",
                "Crores",
            ],
        ]

        for i, row in enumerate(fin_data, 2):
            ws5.append(row)
            for col, val in enumerate(row, 1):
                cell = ws5.cell(i, col)
                cell.border = thin_border
                if "PROFITABILITY" in str(val):
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="27AE60", end_color="27AE60", fill_type="solid"
                    )

        ws5.column_dimensions["A"].width = 25
        ws5.column_dimensions["B"].width = 20
        ws5.column_dimensions["C"].width = 15

        # Sheet 6: Recommendations & Next Steps
        ws6 = wb.create_sheet("Recommendations")

        ws6.cell(1, 1, value="RECOMMENDATIONS")
        ws6.cell(1, 2, value="")
        ws6.cell(1, 1).font = Font(bold=True, size=12, color="FFFFFF")
        ws6.cell(1, 1).fill = PatternFill(
            start_color="27AE60", end_color="27AE60", fill_type="solid"
        )
        ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        for i, rec in enumerate(report.recommendations, 2):
            ws6.cell(i, 1, value=i - 1)
            ws6.cell(i, 2, value=rec)
            ws6.cell(i, 1).font = Font(bold=True)

        row_start = len(report.recommendations) + 4

        ws6.cell(row_start, 1, value="REASONING CHAIN")
        ws6.cell(row_start, 2, value="")
        ws6.cell(row_start, 1).font = Font(bold=True, size=12, color="FFFFFF")
        ws6.cell(row_start, 1).fill = PatternFill(
            start_color="3498DB", end_color="3498DB", fill_type="solid"
        )
        ws6.merge_cells(
            start_row=row_start, start_column=1, end_row=row_start, end_column=2
        )

        for i, step in enumerate(report.reasoning_chain, row_start + 1):
            ws6.cell(i, 1, value=i - row_start)
            ws6.cell(i, 2, value=step)
            ws6.cell(i, 1).font = Font(bold=True)

        row_start = row_start + len(report.reasoning_chain) + 3

        ws6.cell(row_start, 1, value="NEXT STEPS")
        ws6.cell(row_start, 2, value="")
        ws6.cell(row_start, 1).font = Font(bold=True, size=12, color="FFFFFF")
        ws6.cell(row_start, 1).fill = PatternFill(
            start_color="E67E22", end_color="E67E22", fill_type="solid"
        )
        ws6.merge_cells(
            start_row=row_start, start_column=1, end_row=row_start, end_column=2
        )

        for i, step in enumerate(report.next_steps, row_start + 1):
            ws6.cell(i, 1, value=i - row_start)
            ws6.cell(i, 2, value=step)
            ws6.cell(i, 1).font = Font(bold=True)

        ws6.column_dimensions["A"].width = 10
        ws6.column_dimensions["B"].width = 80

        # Save
        wb.save(output_path)
        print(f"  Exported to: {output_path}")

    def export_to_text(self, report: FeasibilityReport, output_path: str):
        """Export report to readable text"""
        lines = []
        lines.append("=" * 70)
        lines.append("FEASIBILITY REPORT - DCPR ANALYSIS")
        lines.append("=" * 70)
        lines.append(f"Report ID: {report.report_id}")
        lines.append(f"Generated: {report.generated_at}")
        lines.append("")

        lines.append("PROPERTY DETAILS")
        lines.append("-" * 40)
        lines.append(f"Survey No: {report.property.survey_no}")
        lines.append(
            f"Area: {report.property.plot_area_sq_m:.0f} sq.m ({report.property.plot_area_sq_ft:.0f} sq.ft)"
        )
        lines.append(f"Road Width: {report.property.road_width_m}m")
        lines.append(f"Zone: {report.property.zone_type}")
        lines.append("")

        lines.append("RECOMMENDED SCHEME")
        lines.append("-" * 40)
        lines.append(f"Scheme: {report.best_scheme}")
        if report.best_scheme in report.all_schemes:
            s = report.all_schemes[report.best_scheme]
            lines.append(f"Total FSI: {s['total_fsi']}")
            lines.append(f"Total BUA: {s['total_bua_sqft']:,.0f} sq.ft")
        lines.append("")

        lines.append("RECOMMENDATIONS")
        lines.append("-" * 40)
        for rec in report.recommendations:
            lines.append(f"- {rec}")
        lines.append("")

        lines.append("REASONING CHAIN")
        lines.append("-" * 40)
        for step in report.reasoning_chain:
            lines.append(step)
        lines.append("")

        lines.append("APPLICABLE DCPR CLAUSES")
        lines.append("-" * 40)
        for clause in report.best_clauses:
            lines.append(f"Clause {clause.clause_id}: {clause.applicability}")
            lines.append(f"  Reason: {clause.reasoning}")
            lines.append("")

        lines.append("NEXT STEPS")
        lines.append("-" * 40)
        for step in report.next_steps:
            lines.append(step)
        lines.append("")

        lines.append("FINANCIAL SUMMARY")
        lines.append("-" * 40)
        for k, v in report.financial_summary.items():
            lines.append(f"{k}: {v}")

        with open(output_path, "w") as f:
            f.write("\n".join(lines))
        print(f"  Exported to: {output_path}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Intelligent Feasibility Analysis")
    parser.add_argument("--survey-no", help="Survey/Plot number")
    parser.add_argument("--area", type=float, help="Plot area in sq.m")
    parser.add_argument("--road-width", type=float, help="Road width in meters")
    parser.add_argument("--zone", default="Residential", help="Zone type")
    parser.add_argument(
        "--affordable", type=float, default=70, help="Affordable housing %"
    )
    parser.add_argument("--file", help="Property card file (PDF/image)")
    parser.add_argument("--output", default="reports/", help="Output directory")

    args = parser.parse_args()

    engine = FeasibilityEngine()
    Path(args.output).mkdir(exist_ok=True)

    if args.file:
        report = engine.analyze_from_file(args.file, args.affordable)
    elif args.survey_no and args.area:
        report = engine.analyze_from_params(
            args.survey_no, args.area, args.road_width or 12, args.zone, args.affordable
        )
    else:
        print("Provide either --file or (--survey-no and --area)")
        exit(1)

    output_base = f"{args.output}{report.report_id}"
    engine.export_to_json(report, f"{output_base}.json")
    engine.export_to_text(report, f"{output_base}.txt")
    engine.export_to_excel(report, f"{output_base}.xlsx")

    print("\nReports saved to:")
    print(f"  - {output_base}.json")
    print(f"  - {output_base}.txt")
    print(f"  - {output_base}.xlsx")
